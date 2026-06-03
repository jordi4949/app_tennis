from fastapi import FastAPI, Request, Form, Depends, HTTPException, status
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.security import HTTPBasic, HTTPBasicCredentials
import secrets
from rapidfuzz import process, fuzz
import unicodedata
import re
import os
from dotenv import load_dotenv
load_dotenv()
from fastapi import UploadFile, File
from openpyxl import load_workbook


from app.database import get_connection

app = FastAPI()
app.mount("/static", StaticFiles(directory="app/Static"), name="static")
templates = Jinja2Templates(directory="app/templates")
security = HTTPBasic()

def comprobar_admin(credentials: HTTPBasicCredentials = Depends(security)):
    ADMIN_USER = os.getenv("ADMIN_USER")
    ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD")

    usuario_correcto = secrets.compare_digest(credentials.username, ADMIN_USER)
    password_correcto = secrets.compare_digest(credentials.password, ADMIN_PASSWORD)

    if not (usuario_correcto and password_correcto):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="No autorizado",
            headers={"WWW-Authenticate": "Basic"},
        )

    return credentials.username

def normalizar_club_para_comparar(texto: str) -> str:
    if not texto:
        return ""

    texto = texto.upper().strip()

    # Quitar acentos
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")

    # Quitar palabras largas que vienen en el Excel federativo
    texto = texto.replace("CLUB DE TENNIS", "")
    texto = texto.replace("CLUB DE TENIS", "")
    texto = texto.replace("CLUB TENNIS", "")
    texto = texto.replace("CLUB TENIS", "")
    texto = texto.replace("CLUB ESPORTIU", "")
    texto = texto.replace("CLUB", "")
    texto = texto.replace("TENNIS", "")
    texto = texto.replace("TENIS", "")

    # Errores típicos OCR
    texto = texto.replace("0", "O")
    texto = texto.replace("1", "I")

    # Separar comas pegadas: "BARCINO,CT" -> "BARCINO CT"
    texto = texto.replace(",", " ")

    # Quitar símbolos raros: puntos, guiones, apóstrofes, etc.
    texto = re.sub(r"[^A-Z0-9 ]", " ", texto)

    # Normalizar abreviaturas frecuentes
    texto = re.sub(r"\bC T\b", "CT", texto)
    texto = re.sub(r"\bT\b", "CT", texto)

    # Quitar siglas finales habituales para comparar por nombre real
    texto = re.sub(r"\bCT\b", "", texto)
    texto = re.sub(r"\bTC\b", "", texto)
    texto = re.sub(r"\bRC\b", "", texto)
    texto = re.sub(r"\bCE\b", "", texto)

    # Espacios repetidos
    texto = re.sub(r"\s+", " ", texto).strip()

    return texto



@app.get("/admin", response_class=HTMLResponse)
def inicio(request: Request, admin: str = Depends(comprobar_admin)):
    return templates.TemplateResponse(
        request=request,
        name="index.html",
        context={"request": request}
    )

@app.get("/admin/jugadores", response_class=HTMLResponse)
def jugadores(
    request: Request,
    buscar: str = "",
    ordenar: str = "apellido",
    admin: str = Depends(comprobar_admin)
):
    texto_busqueda = f"%{buscar.strip()}%"

    if ordenar == "club":
        order_by = "club, apellido1, apellido2, nombre"
    elif ordenar == "licencia":
        order_by = "NULLIF(numero_licencia, '') NULLS LAST, apellido1, apellido2, nombre"
    else:
        order_by = "apellido1, apellido2, nombre"

    conn = get_connection()
    cur = conn.cursor()

    cur.execute(f"""
        SELECT id, nombre, apellido1, apellido2, club, ano_nacimiento, numero_licencia, genero_id
        FROM jugadores
        WHERE
            nombre ILIKE %s
            OR apellido1 ILIKE %s
            OR COALESCE(apellido2, '') ILIKE %s
            OR club ILIKE %s
            OR COALESCE(numero_licencia, '') ILIKE %s
        ORDER BY {order_by}
    """, (texto_busqueda, texto_busqueda, texto_busqueda, texto_busqueda, texto_busqueda))

    jugadores = cur.fetchall()

    cur.execute("""
        SELECT id, nombre
        FROM generos
        ORDER BY id
    """)
    generos = cur.fetchall()

    cur.close()
    conn.close()


    return templates.TemplateResponse(
        request=request,
        name="jugadores.html",
        context={
            "request": request,
            "jugadores": jugadores,
            "buscar": buscar,
            "ordenar": ordenar,
            "generos": generos
        }
    )

@app.post("/admin/jugadores")
def guardar_jugador(
    nombre: str = Form(...),
    apellido1: str = Form(...),
    apellido2: str = Form(""),
    club: str = Form(""),
    ano_nacimiento: int = Form(...),
    numero_licencia: str = Form(""),
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO jugadores (nombre, apellido1, apellido2, club, ano_nacimiento, numero_licencia, genero_id)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (nombre, apellido1, apellido2, club, ano_nacimiento, numero_licencia, genero_id))

    conn.commit()

    cur.close()
    conn.close()

    return RedirectResponse(url="/admin/jugadores", status_code=303)

@app.get("/admin/jugadores/editar/{jugador_id}", response_class=HTMLResponse)
def editar_jugador_form(request: Request, jugador_id: int,
admin: str = Depends(comprobar_admin) 
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, nombre, apellido1, apellido2, club, ano_nacimiento, numero_licencia, genero_id
        FROM jugadores
        WHERE id = %s
    """, (jugador_id,))
    jugador = cur.fetchone()

    cur.execute("""
        SELECT id, nombre
        FROM generos
        ORDER BY id
    """)
    generos = cur.fetchall()

    cur.close()
    conn.close()

    if not jugador:
        return RedirectResponse(url="/admin/jugadores", status_code=303)

    return templates.TemplateResponse(
        request=request,
        name="editar_jugador.html",
        context={
            "request": request,
            "jugador": jugador,
            "generos": generos
        }
    )


@app.post("/admin/jugadores/editar/{jugador_id}")
def actualizar_jugador(
    jugador_id: int,
    nombre: str = Form(...),
    apellido1: str = Form(...),
    apellido2: str = Form(""),
    club: str = Form(""),
    ano_nacimiento: int = Form(...),
    numero_licencia: str = Form(""),
    genero_id: int = Form(...),
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE jugadores
        SET nombre = %s,
            apellido1 = %s,
            apellido2 = %s,
            club = %s,
            ano_nacimiento = %s,
            numero_licencia = %s,
            genero_id = %s
            
        WHERE id = %s
    """, (nombre, apellido1, apellido2, club, ano_nacimiento, numero_licencia, genero_id, jugador_id))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(url="/admin/jugadores", status_code=303)


@app.post("/admin/jugadores/borrar/{jugador_id}")
def borrar_jugador(jugador_id: int,
      admin: str = Depends(comprobar_admin)            
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("DELETE FROM jugadores WHERE id = %s", (jugador_id,))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(url="/admin/jugadores", status_code=303)

def separar_nombre_federacion(texto):
    if not texto or "," not in str(texto):
        return "", "", ""

    apellidos, nombre = str(texto).split(",", 1)

    nombre = nombre.strip().title()
    partes = apellidos.strip().title().split()

    apellido1 = partes[0] if len(partes) >= 1 else ""
    apellido2 = " ".join(partes[1:]) if len(partes) >= 2 else ""

    return nombre, apellido1, apellido2


def obtener_ano_nacimiento(fecha):
    if not fecha:
        return None

    if hasattr(fecha, "year"):
        return fecha.year

    texto = str(fecha).strip()

    if "/" in texto:
        return int(texto.split("/")[-1])

    return None


@app.post("/admin/importar-jugadores/excel-federacion")
async def importar_excel_federacion(
    archivo_excel: UploadFile = File(...),
    genero_id: int = Form(...),
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    wb = load_workbook(archivo_excel.file, data_only=True)
    ws = wb.active

    for fila in ws.iter_rows(min_row=2):
        licencia = fila[1].value
        nombre_completo = fila[2].value
        club = fila[9].value if len(fila) > 9 else ""
        fecha_nacimiento = fila[14].value if len(fila) > 14 else None

        if not licencia or not nombre_completo:
            continue

        nombre, apellido1, apellido2 = separar_nombre_federacion(nombre_completo)
        ano_nacimiento = obtener_ano_nacimiento(fecha_nacimiento)

        cur.execute("""
            INSERT INTO jugadores_importados
            (
                nombre,
                apellido1,
                apellido2,
                club,
                ano_nacimiento,
                numero_licencia,
                genero_id
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            nombre,
            apellido1,
            apellido2,
            club,
            ano_nacimiento,
            str(licencia).strip(),
            genero_id
        ))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(url="/admin/importar-jugadores", status_code=303)



@app.get("/admin/importar-jugadores")
def ver_importados(
    request: Request,
    buscar: str = "",
    ordenar: str = "apellido",
    genero_id: int = 0,
    admin: str = Depends(comprobar_admin)
):
    if ordenar == "club":
        order_by = "club, apellido1, apellido2, nombre"
    elif ordenar == "licencia":
        order_by = "NULLIF(numero_licencia, '') NULLS LAST, apellido1, apellido2, nombre"
    elif ordenar == "genero":
        order_by = "genero_id, apellido1, apellido2, nombre"
    else:
        order_by = "apellido1, apellido2, nombre"

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, nombre
        FROM generos
        ORDER BY id
    """)
    generos = cur.fetchall()

    condiciones = []
    parametros = []

    if buscar:
        texto_busqueda = f"%{buscar.strip()}%"
        condiciones.append("""
            (
                nombre ILIKE %s
                OR apellido1 ILIKE %s
                OR COALESCE(apellido2, '') ILIKE %s
                OR club ILIKE %s
                OR COALESCE(numero_licencia, '') ILIKE %s
            )
        """)
        parametros.extend([
            texto_busqueda,
            texto_busqueda,
            texto_busqueda,
            texto_busqueda,
            texto_busqueda
        ])

    if genero_id != 0:
        condiciones.append("genero_id = %s")
        parametros.append(genero_id)

    where_sql = ""

    if condiciones:
        where_sql = "WHERE " + " AND ".join(condiciones)

    cur.execute(f"""
        SELECT
            id,
            nombre,
            apellido1,
            apellido2,
            club,
            ano_nacimiento,
            numero_licencia,
            genero_id
        FROM jugadores_importados
        {where_sql}
        ORDER BY {order_by}
    """, parametros)

    jugadores = cur.fetchall()

    cur.close()
    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="importar_jugadores.html",
        context={
            "request": request,
            "jugadores": jugadores,
            "buscar": buscar,
            "ordenar": ordenar,
            "genero_id": genero_id,
            "generos": generos
        }
    )

def corregir_club_alias_manual(club: str) -> str:
    if not club:
        return club

    clave = normalizar_club_para_comparar(club)

    alias = {
        "ANDRES GIMENO": "Andrés Gimeno, CT",
        "BARA": "Barà, CT",
        "BLANES": "Blanes, CT",
        "CARDEDEU": "Cardedeu, CT",
        "DARO": "D'Aro, CT",
        "EL ROMANI": "El Romaní, CT",
        "LLAFRANC": "Llafranc, CT",
        "MANLLEU": "Manlleu, CT",
        "MATARO": "Mataró, CT",
        "MONTBLANC": "Montblanc, CT",
        "OLESA": "Olesa, CT",
        "PINEDA GAVA": "Pineda Gavà, CT",
        "RIPOLLET": "Ripollet, CT",
        "ROSES": "Roses, CT",
        "SABADELL": "Sabadell, CT",
        "SANT CELONI": "Sant Celoni, CT",
        "SERRAMAR": "Serramar, CT",
        "TARREGA": "Tàrrega, CT",
        "CERDANYOLA": "Cerdanyola, CT",
        "GIRONA": "Girona, CT",
        "BISBAL EMPORDA": "La Bisbal d'Empordà, CT",
        "PALLEJA": "Pallejà, CT",
        "SANT BOI": "Sant Boi, CT",
        "URGELL": "Urgell, CT",
        "COSTA BRAVA": "Costa Brava, CT",
        "ELS GORCHS": "Els Gorchs, CT",
        "POBLA CLARAMUNT": "La Pobla de Claramunt, CT",
        "VIC": "Vic, CT",
        "VILANOVA": "Vilanova, CT",
        "BARCINO": "Barcino, CT",
        "SALUT 1902": "La Salut 1902, CT",
        "FIGUERES": "Figueres, CT",
        "METLLA MAR": "L'Ametlla de Mar, CT",
        "LLEIDA": "Lleida, CT",
        "NATACIO SANT CUGAT": "Natació Sant Cugat, CT",
        "PINEDA": "Pineda, CT",
        "PORQUERES": "Porqueres, CT",
        "REUS MONTEROLS": "Reus Monterols, CT",
        "SITGES": "Sitges, CT",
        "TARRAGONA": "Tarragona, CT",
        "TORELLO": "Torelló, CT",
        "TORREDEMBARRA": "Torredembarra, CT",
        "TORTOSA": "Tortosa, CT",
        "BERGA": "Berga, TC",
        "BADALONA": "Badalona, TC",
        "BARCELONA 1899": "Barcelona-1899, RCT",
        "POLO": "Real Club de Polo",
        "TOPTEN": "Topten Tennis Club",
        "VALL HEBRON": "Vall d'Hebron",
        "EGARA": "Egara, Club",
        "FARNERS": "Farners Tennis Club",
        "FEDERACIO CATALANA": "Federació Catalana de Tennis",
    }

    return alias.get(clave, club)


@app.post("/admin/importar-jugadores/corregir-clubs")
def corregir_clubs_importados(admin: str = Depends(comprobar_admin)):
    conn = get_connection()
    cur = conn.cursor()

    # Clubs distintos que vienen del OCR
    cur.execute("""
        SELECT DISTINCT club
        FROM jugadores_importados
        WHERE club IS NOT NULL
          AND TRIM(club) <> ''
    """)
    clubs_importados = [fila[0] for fila in cur.fetchall()]

    # Clubs buenos desde tabla clubs
    cur.execute("""
        SELECT nombre
        FROM clubs
        WHERE nombre IS NOT NULL
          AND TRIM(nombre) <> ''
    """)
    clubs_tabla = [fila[0] for fila in cur.fetchall()]

    # Clubs buenos desde jugadores ya aprobados
    cur.execute("""
        SELECT DISTINCT club
        FROM jugadores
        WHERE club IS NOT NULL
          AND TRIM(club) <> ''
    """)
    clubs_jugadores = [fila[0] for fila in cur.fetchall()]

    clubs_buenos = sorted(set(clubs_tabla + clubs_jugadores))

    if not clubs_importados or not clubs_buenos:
        cur.close()
        conn.close()
        return RedirectResponse(
            url="/admin/importar-jugadores?ordenar=club",
            status_code=303
        )

    clubs_buenos_normalizados = {
        normalizar_club_para_comparar(club): club
        for club in clubs_buenos
    }

    for club_importado in clubs_importados:

        club_alias = corregir_club_alias_manual(club_importado)

        if club_alias != club_importado:
            cur.execute("""
                UPDATE jugadores_importados
                SET club = %s
                WHERE club = %s
            """, (club_alias, club_importado))
            continue

        club_importado_norm = normalizar_club_para_comparar(club_importado)
    
    

        mejor = process.extractOne(
            club_importado_norm,
            list(clubs_buenos_normalizados.keys()),
            scorer=fuzz.token_set_ratio
        )


        if not mejor:
            continue

        club_bueno_norm, puntuacion, _ = mejor
        club_bueno = clubs_buenos_normalizados[club_bueno_norm]

        print("CLUB IMPORTADO:", club_importado, "=>", club_importado_norm)
        print("MEJOR:", club_bueno, "=>", club_bueno_norm, "PUNTOS:", puntuacion)


        # Umbral alto para evitar cambios peligrosos.
        # Si corrige poco, luego bajamos a 85.
        if puntuacion >= 75 and club_importado != club_bueno:
            cur.execute("""
                UPDATE jugadores_importados
                SET club = %s
                WHERE club = %s
            """, (club_bueno, club_importado))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(
        url="/admin/importar-jugadores?ordenar=club",
        status_code=303
    )

    
@app.get("/admin/importar-jugadores/aprobar/{jugador_id}")
def aprobar_jugador_importado(
    jugador_id: int,
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT nombre, apellido1, apellido2, club, ano_nacimiento, numero_licencia
        FROM jugadores_importados
        WHERE id = %s
    """, (jugador_id,))

    jugador = cur.fetchone()

    if jugador:
        cur.execute("""
            INSERT INTO clubs (nombre)
            VALUES (%s)
            ON CONFLICT (nombre) DO NOTHING
        """, (jugador[3],))
        
        cur.execute("""
            INSERT INTO jugadores
            (nombre, apellido1, apellido2, club, ano_nacimiento, numero_licencia)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (numero_licencia) DO NOTHING
        """, jugador)

        cur.execute("""
            DELETE FROM jugadores_importados
            WHERE id = %s
        """, (jugador_id,))

        conn.commit()

    cur.close()
    conn.close()

    return RedirectResponse(url="/admin/importar-jugadores", status_code=303)

@app.get("/admin/importar-jugadores/borrar/{jugador_id}")
def borrar_importado(
    jugador_id: int,
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        DELETE FROM jugadores_importados
        WHERE id = %s
    """, (jugador_id,))

    conn.commit()

    cur.close()
    conn.close()

    return RedirectResponse(url="/admin/importar-jugadores", status_code=303)

@app.get("/admin/importar-jugadores/editar/{jugador_id}")
def editar_importado(
    request: Request,
    jugador_id: int,
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    # jugador
    cur.execute("""
        SELECT id, nombre, apellido1, apellido2, club, ano_nacimiento, numero_licencia
        FROM jugadores_importados
        WHERE id = %s
    """, (jugador_id,))
    jugador = cur.fetchone()

    # lista de clubs existentes
    cur.execute("""
    SELECT nombre
    FROM clubs
    ORDER BY nombre ASC
""")
    
    clubs = cur.fetchall()

    cur.close()
    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="editar_jugador_importado.html",
        context={
            "request": request,
            "jugador": jugador,
            "clubs": clubs
        }
    )
@app.post("/admin/importar-jugadores/editar/{jugador_id}")
def guardar_importado(
    jugador_id: int,
    nombre: str = Form(...),
    apellido1: str = Form(...),
    apellido2: str = Form(""),
    club_existente: str = Form(""),
    club_nuevo: str = Form(""),
    ano_nacimiento: int = Form(...),
    numero_licencia: str = Form(...),
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    club_final = club_nuevo.strip() if club_nuevo.strip() else club_existente
    cur.execute("""
        INSERT INTO clubs (nombre)
        VALUES (%s)
        ON CONFLICT (nombre) DO NOTHING
    """, (club_final,))


    cur.execute("""
        UPDATE jugadores_importados
        SET nombre=%s, apellido1=%s, apellido2=%s,
            club=%s, ano_nacimiento=%s, numero_licencia=%s
        WHERE id=%s
    """, (
        nombre, apellido1, apellido2,
        club_final, ano_nacimiento, numero_licencia,
        jugador_id
    ))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse("/admin/importar-jugadores", status_code=303)

@app.post("/admin/importar-jugadores/aprobar-seleccionados")
def aprobar_seleccionados(
    jugadores_ids: list[int] = Form(...),
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    for jugador_id in jugadores_ids:

        cur.execute("""
            SELECT nombre, apellido1, apellido2, club, ano_nacimiento, numero_licencia
            FROM jugadores_importados
            WHERE id = %s
        """, (jugador_id,))

        jugador = cur.fetchone()

        if jugador:
            # Insertar club
            cur.execute("""
                INSERT INTO clubs (nombre)
                VALUES (%s)
                ON CONFLICT (nombre) DO NOTHING
            """, (jugador[3],))

            # Insertar jugador
            cur.execute("""
                INSERT INTO jugadores
                (nombre, apellido1, apellido2, club, ano_nacimiento, numero_licencia)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (numero_licencia) DO NOTHING
            """, jugador)

            # Borrar importado
            cur.execute("""
                DELETE FROM jugadores_importados
                WHERE id = %s
            """, (jugador_id,))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse("/admin/importar-jugadores", status_code=303)

@app.get("/admin/torneos", response_class=HTMLResponse)
def ver_torneos(request: Request, admin: str = Depends(comprobar_admin)):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, nombre, fecha_inicio, categoria, ubicacion
        FROM torneos
        ORDER BY id;
    """)
    torneos = cur.fetchall()

    cur.close()
    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="torneos.html",
        context={
            "request": request,
            "torneos": torneos
        }
    )


@app.post("/admin/torneos")
def guardar_torneo(
    nombre: str = Form(...),
    fecha_inicio: str = Form(...),
    categoria: str = Form(...),
    ubicacion: str = Form(...),
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO torneos (nombre, fecha_inicio, categoria, ubicacion)
        VALUES (%s, %s, %s, %s)
    """, (nombre, fecha_inicio, categoria, ubicacion))

    conn.commit()

    cur.close()
    conn.close()

    return RedirectResponse(url="/admin/torneos", status_code=303)
@app.get("/admin/torneos/editar/{torneo_id}", response_class=HTMLResponse)
def editar_torneo_form(request: Request, torneo_id: int,
admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, nombre, fecha_inicio, categoria, ubicacion
        FROM torneos
        WHERE id = %s
    """, (torneo_id,))

    torneo = cur.fetchone()

    cur.close()
    conn.close()

    if not torneo:
        return RedirectResponse(url="/admin/torneos", status_code=303)

    return templates.TemplateResponse(
        request=request,
        name="editar_torneo.html",
        context={
            "request": request,
            "torneo": torneo
        }
    )


@app.post("/admin/torneos/editar/{torneo_id}")
def actualizar_torneo(
    torneo_id: int,
    nombre: str = Form(...),
    fecha_inicio: str = Form(...),
    categoria: str = Form(...),
    ubicacion: str = Form(...),
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE torneos
        SET nombre = %s,
            fecha_inicio = %s,
            categoria = %s,
            ubicacion = %s
        WHERE id = %s
    """, (nombre, fecha_inicio, categoria, ubicacion, torneo_id))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(url="/admin/torneos", status_code=303)


@app.post("/admin/torneos/borrar/{torneo_id}")
def borrar_torneo(torneo_id: int,
admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("DELETE FROM torneos WHERE id = %s", (torneo_id,))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(url="/admin/torneos", status_code=303)

@app.get("/admin/cuadros", response_class=HTMLResponse)
def ver_cuadros(
    request: Request,
    torneo_id: int = 0,
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, nombre, fecha_inicio, categoria, ubicacion
        FROM torneos
        ORDER BY fecha_inicio DESC, nombre
    """)
    torneos = cur.fetchall()

    cur.execute("""
        SELECT id, nombre
        FROM generos
        ORDER BY id
    """)
    generos = cur.fetchall()

    cur.execute("""
        SELECT id, nombre
        FROM categorias
        ORDER BY id
    """)
    categorias = cur.fetchall()

    if torneo_id != 0:
        cur.execute("""
            SELECT
                c.id,
                c.nombre,
                c.tamano,
                c.numero_jugadores,
                COALESCE(c.observaciones, ''),
                t.nombre,
                t.fecha_inicio,
                COALESCE(cat.nombre, ''),
                COALESCE(g.nombre, ''),
                t.ubicacion,
                
                (
                    SELECT COUNT(*)
                    FROM rondas_cuadro rc
                    WHERE rc.cuadro_id = c.id
                        AND rc.ganador_id IS NOT NULL
                ) AS resultados_guardados,

                (
                    SELECT TRIM(
                        j.nombre || ' ' ||
                        j.apellido1 || ' ' ||
                        COALESCE(j.apellido2, '')
                    )
                
                    FROM rondas_cuadro rc
                    JOIN jugadores j ON j.id = rc.ganador_id
                    WHERE rc.cuadro_id = c.id
                    ORDER BY rc.ronda_numero DESC,
                            rc.posicion_ronda DESC
                    LIMIT 1

                ) AS ganador    

                    

            FROM cuadros c
            JOIN torneos t ON c.torneo_id = t.id
            LEFT JOIN categorias cat ON c.categoria_id = cat.id
            LEFT JOIN generos g ON c.genero_id = g.id
            WHERE c.torneo_id = %s
            ORDER BY cat.nombre, g.nombre, c.nombre
        """, (torneo_id,))
    else:
        cur.execute("""
            SELECT
                c.id,
                c.nombre,
                c.tamano,
                c.numero_jugadores,
                COALESCE(c.observaciones, ''),
                t.nombre,
                t.fecha_inicio,
                COALESCE(cat.nombre, ''),
                COALESCE(g.nombre, ''),
                t.ubicacion,

                (
                    SELECT COUNT(*)
                    FROM rondas_cuadro rc
                    WHERE rc.cuadro_id = c.id
                        AND rc.ganador_id IS NOT NULL
                ) AS resultados_guardados,

                (
                    SELECT TRIM(
                        j.nombre || ' ' ||
                        j.apellido1 || ' ' ||
                        COALESCE(j.apellido2, '')
                    )
                
                    FROM rondas_cuadro rc
                    JOIN jugadores j ON j.id = rc.ganador_id
                    WHERE rc.cuadro_id = c.id
                    ORDER BY rc.ronda_numero DESC,
                            rc.posicion_ronda DESC
                    LIMIT 1

                ) AS ganador
                    
            FROM cuadros c
            JOIN torneos t ON c.torneo_id = t.id
            LEFT JOIN categorias cat ON c.categoria_id = cat.id
            LEFT JOIN generos g ON c.genero_id = g.id
            ORDER BY t.fecha_inicio DESC, t.nombre, cat.nombre, g.nombre, c.nombre
        """)   

    cuadros = cur.fetchall()

    cur.close()
    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="cuadros.html",
        context={
            "request": request,
            "torneos": torneos,
            "cuadros": cuadros,
            "torneo_id": torneo_id,
            "generos": generos,
            "categorias": categorias
        }
    )

@app.post("/admin/cuadros/{cuadro_id}/importar-excel-archivo")
def importar_excel_archivo(
    cuadro_id: int,
    file: UploadFile = File(...),
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    wb = load_workbook(file.file)
    ws = wb.active

    # Borrar inscritos anteriores del cuadro para no duplicar
    cur.execute("""
        DELETE FROM cuadro_inscritos
        WHERE cuadro_id = %s
    """, (cuadro_id,))

    for fila_excel in ws.iter_rows(min_row=2):
        licencia = fila_excel[1].value  # columna B
        nombre_excel = fila_excel[2].value  # columna C

        if not licencia:
            continue

        licencia = str(licencia).strip().replace(".0", "")

        cur.execute("""
            SELECT id
            FROM jugadores
            WHERE TRIM(numero_licencia) = %s
        """, (licencia,))

        jugador = cur.fetchone()

        if jugador:
            jugador_id = jugador[0]
            estado = "encontrado"
        else:
            jugador_id = None
            estado = "no_encontrado"

        cur.execute("""
            INSERT INTO cuadro_inscritos
            (cuadro_id, jugador_id, numero_licencia, nombre_excel, estado)
            VALUES (%s, %s, %s, %s, %s)
        """, (cuadro_id, jugador_id, licencia, nombre_excel, estado))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(
        url=f"/admin/cuadros/{cuadro_id}/inscritos",
        status_code=303
    )
@app.get("/admin/cuadros/{cuadro_id}/inscritos", response_class=HTMLResponse)
def ver_inscritos(
    cuadro_id: int,
    request: Request,
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            ci.id,
            ci.numero_licencia,
            ci.nombre_excel,
            j.id,
            j.nombre,
            j.apellido1,
            j.apellido2,
            ci.estado,
            ci.posicion
        FROM cuadro_inscritos ci
        LEFT JOIN jugadores j ON ci.jugador_id = j.id
        WHERE ci.cuadro_id = %s
        ORDER BY ci.posicion NULLS LAST, ci.nombre_excel
    """, (cuadro_id,))

    inscritos = cur.fetchall()

    cur.execute("""
        SELECT tamano
        FROM cuadros
        WHERE id = %s
    """, (cuadro_id,))

    cuadro = cur.fetchone()
    tamano_cuadro = cuadro[0] if cuadro else 32
    posiciones = list(range(1, tamano_cuadro + 1))

    posiciones_ocupadas = [
        i[8] for i in inscritos
        if i[8] is not None
    ]

    inscritos_por_posicion = {
        i[8]: i
        for i in inscritos
        if i[8] is not None
    }

    cuadro_ordenado = []

    for posicion in posiciones:
        inscrito = inscritos_por_posicion.get(posicion)

        if inscrito:
            cuadro_ordenado.append({
                "posicion": posicion,
                "tipo": "jugador",
                "inscrito": inscrito
            })
        else:
            cuadro_ordenado.append({
                "posicion": posicion,
                "tipo": "bye",
                "inscrito": None
            })

    cur.close()
    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="inscritos.html",
        context={
            "request": request,
            "inscritos": inscritos,
            "cuadro_id": cuadro_id,
            "posiciones": posiciones,
            "posiciones_ocupadas": posiciones_ocupadas,
            "cuadro_ordenado": cuadro_ordenado
        }
    )

@app.get("/admin/cuadros/{cuadro_id}/resultados", response_class=HTMLResponse)
def resultados_cuadro(
    cuadro_id: int,
    request: Request,
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            ci.id,
            ci.numero_licencia,
            ci.nombre_excel,
            j.id,
            j.nombre,
            j.apellido1,
            j.apellido2,
            ci.estado,
            ci.posicion
        FROM cuadro_inscritos ci
        LEFT JOIN jugadores j ON ci.jugador_id = j.id
        WHERE ci.cuadro_id = %s
        ORDER BY ci.posicion NULLS LAST, ci.nombre_excel
    """, (cuadro_id,))

    inscritos = cur.fetchall()

    cur.execute("""
        SELECT tamano
        FROM cuadros
        WHERE id = %s
    """, (cuadro_id,))

    cuadro = cur.fetchone()
    tamano_cuadro = cuadro[0] if cuadro else 32
    posiciones = list(range(1, tamano_cuadro + 1))

    inscritos_por_posicion = {
        i[8]: i
        for i in inscritos
        if i[8] is not None
    }

    cuadro_ordenado = []

    for posicion in posiciones:
        inscrito = inscritos_por_posicion.get(posicion)

        if inscrito:
            cuadro_ordenado.append({
                "posicion": posicion,
                "tipo": "jugador",
                "inscrito": inscrito
            })
        else:
            cuadro_ordenado.append({
                "posicion": posicion,
                "tipo": "bye",
                "inscrito": None
            })

    cur.execute("""
        SELECT
            p.posicion_ronda,
            p.estado,
            p.resultado,
            p.id
        FROM partidos p
        WHERE p.cuadro_id = %s
          AND p.ronda_numero = 1
    """, (cuadro_id,))

    partidos_guardados = {
        fila[0]: {
            "estado": fila[1],
            "resultado": fila[2],
            "partido_id": fila[3]
        }
        for fila in cur.fetchall()
    }

    emparejamientos = []
    numero_partido = 1

    for i in range(0, len(cuadro_ordenado), 2):
        lado1 = cuadro_ordenado[i]
        lado2 = cuadro_ordenado[i + 1] if i + 1 < len(cuadro_ordenado) else {
            "posicion": None,
            "tipo": "bye",
            "inscrito": None
        }

        if lado1["tipo"] == "jugador" and lado2["tipo"] == "jugador":
            estado = "partido"
        elif lado1["tipo"] == "jugador" and lado2["tipo"] == "bye":
            estado = "bye_jugador1"
        elif lado1["tipo"] == "bye" and lado2["tipo"] == "jugador":
            estado = "bye_jugador2"
        else:
            estado = "vacio"

        guardado = partidos_guardados.get(numero_partido)

        emparejamientos.append({
            "numero_partido": numero_partido,
            "lado1": lado1,
            "lado2": lado2,
            "estado": estado,
            "guardado": guardado
        })

        numero_partido += 1

    partidos_ronda_1 = []

    for partido in emparejamientos:
        jugador1_id = None
        jugador2_id = None
        jugador1_nombre = "BYE"
        jugador2_nombre = "BYE"


        if partido["lado1"]["tipo"] == "jugador":
            inscrito1 = partido["lado1"]["inscrito"]

            jugador1_id = inscrito1[3]

            jugador1_nombre = (
                f"{inscrito1[4]} "
                f"{inscrito1[5]} "
                f"{inscrito1[6] or ''}"
            ).strip()

        if partido["lado2"]["tipo"] == "jugador":
            inscrito2 = partido["lado2"]["inscrito"]

            jugador2_id = inscrito2[3]

            jugador2_nombre = (
                f"{inscrito2[4]} "
                f"{inscrito2[5]} "
                f"{inscrito2[6] or ''}"
            ).strip()


        partidos_ronda_1.append({
            "numero_partido": partido["numero_partido"],
            "jugador1_id": jugador1_id,
            "jugador1_nombre": jugador1_nombre,
            "jugador2_id": jugador2_id,
            "jugador2_nombre": jugador2_nombre,
            "ganador_id": None,
            "resultado": partido["guardado"]["resultado"] if partido["guardado"] else None,
            "estado": partido["estado"]
        })

    nombres_rondas = nombres_rondas_por_tamano(tamano_cuadro)
    rondas = []

    for numero_ronda, nombre_ronda in enumerate(nombres_rondas, start=1):
        prefijo = "" if numero_ronda == 1 else f"r{numero_ronda}_"

        if numero_ronda == 1:
            rondas.append({
                "numero": 1,
                "nombre": nombre_ronda,
                "prefijo": prefijo,
                "es_final": len(nombres_rondas) == 1,
                "partidos": partidos_ronda_1
            })
            continue

        cur.execute("""
            SELECT
                rc.posicion_ronda,
                rc.jugador1_id,
                TRIM(j1.nombre || ' ' || j1.apellido1 || ' ' || COALESCE(j1.apellido2, '')) AS jugador1_nombre,
                rc.jugador2_id,
                TRIM(j2.nombre || ' ' || j2.apellido1 || ' ' || COALESCE(j2.apellido2, '')) AS jugador2_nombre,
                rc.ganador_id,
                rc.resultado,
                rc.estado
            FROM rondas_cuadro rc
            LEFT JOIN jugadores j1 ON rc.jugador1_id = j1.id
            LEFT JOIN jugadores j2 ON rc.jugador2_id = j2.id
            WHERE rc.cuadro_id = %s
              AND rc.ronda_numero = %s
            ORDER BY rc.posicion_ronda
        """, (cuadro_id, numero_ronda))

        partidos = []

        for fila in cur.fetchall():
            partidos.append({
                "numero_partido": fila[0],
                "jugador1_id": fila[1],
                "jugador1_nombre": fila[2],
                "jugador2_id": fila[3],
                "jugador2_nombre": fila[4],
                "ganador_id": fila[5],
                "resultado": fila[6],
                "estado": fila[7]
            })

        if partidos:
            rondas.append({
                "numero": numero_ronda,
                "nombre": nombre_ronda,
                "prefijo": prefijo,
                "es_final": numero_ronda == len(nombres_rondas),
                "partidos": partidos
            })

    cur.close()
    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="resultados_cuadro.html",
        context={
            "request": request,
            "cuadro_id": cuadro_id,
            "emparejamientos": emparejamientos,
            "rondas": rondas
        }
    )

def guardar_o_actualizar_ronda_cuadro(
    cur,
    cuadro_id,
    ronda_numero,
    nombre_ronda,
    posicion_ronda,
    jugador1_id,
    jugador2_id,
    ganador_id,
    jugador1_posicion,
    jugador2_posicion,
    estado,
    resultado,
    partido_id=None
):
    cur.execute("""
        INSERT INTO rondas_cuadro
        (
            cuadro_id,
            ronda_numero,
            nombre_ronda,
            posicion_ronda,
            jugador1_id,
            jugador2_id,
            ganador_id,
            jugador1_posicion,
            jugador2_posicion,
            estado,
            resultado,
            partido_id
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (cuadro_id, ronda_numero, posicion_ronda)
        DO UPDATE SET
            nombre_ronda = EXCLUDED.nombre_ronda,
            jugador1_id = EXCLUDED.jugador1_id,
            jugador2_id = EXCLUDED.jugador2_id,
            ganador_id = EXCLUDED.ganador_id,
            jugador1_posicion = EXCLUDED.jugador1_posicion,
            jugador2_posicion = EXCLUDED.jugador2_posicion,
            estado = EXCLUDED.estado,
            resultado = EXCLUDED.resultado,
            partido_id = EXCLUDED.partido_id
        RETURNING id
    """, (
        cuadro_id,
        ronda_numero,
        nombre_ronda,
        posicion_ronda,
        jugador1_id,
        jugador2_id,
        ganador_id,
        jugador1_posicion,
        jugador2_posicion,
        estado,
        resultado,
        partido_id
    ))

    return cur.fetchone()[0]

def nombres_rondas_por_tamano(tamano):

    if tamano == 8:
        return [
            "Cuartos",
            "Semifinal",
            "Final"
        ]

    elif tamano == 16:
        return [
            "Octavos",
            "Cuartos",
            "Semifinal",
            "Final"
        ]

    elif tamano == 32:
        return [
            "Dieciseisavos",
            "Octavos",
            "Cuartos",
            "Semifinal",
            "Final"
        ]

    elif tamano == 64:
        return [
            "Treintaidosavos",
            "Dieciseisavos",
            "Octavos",
            "Cuartos",
            "Semifinal",
            "Final"
        ]

    elif tamano == 128:
        return [
            "Sesentaicuatroavos",
            "Treintaidosavos",
            "Dieciseisavos",
            "Octavos",
            "Cuartos",
            "Semifinal",
            "Final"
        ]

    return []

def generar_siguiente_ronda(
    cur,
    cuadro_id,
    ronda_actual,
    nombre_siguiente_ronda
):
    ronda_siguiente = ronda_actual + 1

    cur.execute("""
        SELECT
            rc.posicion_ronda,
            COALESCE(rc.ganador_id, p.ganador_id) AS ganador_id
        FROM rondas_cuadro rc
        LEFT JOIN partidos p ON p.id = rc.partido_id
        WHERE rc.cuadro_id = %s
          AND rc.ronda_numero = %s
        ORDER BY rc.posicion_ronda
    """, (cuadro_id, ronda_actual))

    partidos_ronda = cur.fetchall()

    for i in range(0, len(partidos_ronda), 2):

        if i + 1 >= len(partidos_ronda):
            break

        ganador1 = partidos_ronda[i][1]
        ganador2 = partidos_ronda[i + 1][1]

        if not ganador1 or not ganador2:
            continue

        numero_partido = (i // 2) + 1

        cur.execute("""
            INSERT INTO rondas_cuadro
            (
                cuadro_id,
                ronda_numero,
                nombre_ronda,
                posicion_ronda,
                jugador1_id,
                jugador2_id,
                estado
            )
            VALUES (%s, %s, %s, %s, %s, %s, 'pendiente')
            ON CONFLICT (cuadro_id, ronda_numero, posicion_ronda)
            DO UPDATE SET
                jugador1_id = EXCLUDED.jugador1_id,
                jugador2_id = EXCLUDED.jugador2_id,
                estado = 'pendiente'
        """, (
            cuadro_id,
            ronda_siguiente,
            nombre_siguiente_ronda,
            numero_partido,
            ganador1,
            ganador2
        ))


def ganador_set(j1, j2):
    if (j1 == 6 and j2 <= 4) or (j1 == 7 and j2 in [5, 6]):
        return 1
    if (j2 == 6 and j1 <= 4) or (j2 == 7 and j1 in [5, 6]):
        return 2
    return 0


def tiebreak_valido(j1, j2, minimo):
    ganador = max(j1, j2)
    perdedor = min(j1, j2)
    return ganador >= minimo and ganador - perdedor >= 2

def guardar_o_actualizar_bye(
    cur,
    torneo_id,
    cuadro_id,
    nombre_primera_ronda,
    numero_partido,
    jugador1_id,
    jugador2_id,
    ganador_id,
    jugador1_pos,
    jugador2_pos
):
    ronda_cuadro_id = guardar_o_actualizar_ronda_cuadro(
        cur,
        cuadro_id,
        1,
        nombre_primera_ronda,
        numero_partido,
        jugador1_id,
        jugador2_id,
        ganador_id,
        jugador1_pos,
        jugador2_pos,
        "bye",
        "BYE",
        None
    )

    cur.execute("""
        SELECT id
        FROM partidos
        WHERE cuadro_id = %s
          AND ronda_numero = 1
          AND posicion_ronda = %s
    """, (cuadro_id, numero_partido))

    partido_existente = cur.fetchone()

    if partido_existente:
        partido_id = partido_existente[0]

        cur.execute("""
            UPDATE partidos
            SET jugador1_id = %s,
                jugador2_id = %s,
                ganador_id = %s,
                ronda = %s,
                resultado = %s,
                jugador1_posicion = %s,
                jugador2_posicion = %s,
                estado = 'bye',
                ronda_cuadro_id = %s
            WHERE id = %s
        """, (
            jugador1_id,
            jugador2_id,
            ganador_id,
            nombre_primera_ronda,
            "BYE",
            jugador1_pos,
            jugador2_pos,
            ronda_cuadro_id,
            partido_id
        ))

        cur.execute("""
            DELETE FROM sets
            WHERE partido_id = %s
        """, (partido_id,))

    else:
        cur.execute("""
            INSERT INTO partidos
            (
                torneo_id,
                fecha_partido,
                jugador1_id,
                jugador2_id,
                ganador_id,
                ronda,
                resultado,
                cuadro_id,
                ronda_numero,
                posicion_ronda,
                jugador1_posicion,
                jugador2_posicion,
                estado,
                ronda_cuadro_id
            )
            VALUES (%s, CURRENT_DATE, %s, %s, %s, %s, %s, %s, 1, %s, %s, %s, 'bye', %s)
            RETURNING id
        """, (
            torneo_id,
            jugador1_id,
            jugador2_id,
            ganador_id,
            nombre_primera_ronda,
            "BYE",
            cuadro_id,
            numero_partido,
            jugador1_pos,
            jugador2_pos,
            ronda_cuadro_id
        ))
        partido_id = cur.fetchone()[0]

    cur.execute("""
        UPDATE rondas_cuadro
        SET partido_id = %s
        WHERE id = %s
    """, (partido_id, ronda_cuadro_id))    


@app.post("/admin/cuadros/{cuadro_id}/guardar-resultados")
async def guardar_resultados_cuadro(
    cuadro_id: int,
    request: Request,
    admin: str = Depends(comprobar_admin)
):
    form = await request.form()
    
    print("FORM RECIBIDO:", dict(form))

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT torneo_id
        FROM cuadros
        WHERE id = %s
    """, (cuadro_id,))
    fila_cuadro = cur.fetchone()

    if not fila_cuadro:
        cur.close()
        conn.close()
        return RedirectResponse(url="/admin/cuadros", status_code=303)
    


    torneo_id = fila_cuadro[0]

    cur.execute("""
        SELECT
            ci.id,
            j.id,
            ci.posicion
        FROM cuadro_inscritos ci
        LEFT JOIN jugadores j ON ci.jugador_id = j.id
        WHERE ci.cuadro_id = %s
          AND ci.posicion IS NOT NULL
        ORDER BY ci.posicion
    """, (cuadro_id,))

    inscritos_posiciones = cur.fetchall()

    jugadores_por_posicion = {
        fila[2]: fila[1]
        for fila in inscritos_posiciones
    }

    cur.execute("""
        SELECT tamano
        FROM cuadros
        WHERE id = %s
    """, (cuadro_id,))

    fila_tamano = cur.fetchone()
    tamano_cuadro = fila_tamano[0] if fila_tamano else 32
    nombres_rondas = nombres_rondas_por_tamano(tamano_cuadro)

    nombre_primera_ronda = nombres_rondas[0] if nombres_rondas else "Primera ronda"
    numero_partido_bye = 1

    for posicion in range(1, tamano_cuadro + 1, 2):
        pos1 = posicion
        pos2 = posicion + 1

        jugador1_id = jugadores_por_posicion.get(pos1)
        jugador2_id = jugadores_por_posicion.get(pos2)

        if jugador1_id and not jugador2_id:
            print("GUARDANDO BYE J1:", numero_partido_bye, pos1, pos2, jugador1_id)    
            guardar_o_actualizar_bye(
                cur,
                torneo_id,
                cuadro_id,
                nombre_primera_ronda,
                numero_partido_bye,
                jugador1_id,
                None,
                jugador1_id,
                pos1,
                pos2
            )

        elif jugador2_id and not jugador1_id:
            print("GUARDANDO BYE J2:", numero_partido_bye, pos1, pos2, jugador2_id)
            guardar_o_actualizar_bye(
                cur,
                torneo_id,
                cuadro_id,
                nombre_primera_ronda,
                numero_partido_bye,
                None,
                jugador2_id,
                jugador2_id,
                pos1,
                pos2
            )

        numero_partido_bye += 1

    numeros_partido = []

    for key in form.keys():
        if key.startswith("jugador1_id_") and not key.startswith("jugador1_id_r"):
            numeros_partido.append(int(key.replace("jugador1_id_", "")))

    for numero_partido in numeros_partido:
        jugador1_id = int(form.get(f"jugador1_id_{numero_partido}"))
        jugador2_id = int(form.get(f"jugador2_id_{numero_partido}"))
        jugador1_pos = int(form.get(f"jugador1_pos_{numero_partido}"))
        jugador2_pos = int(form.get(f"jugador2_pos_{numero_partido}"))

        set1_j1 = int(form.get(f"set1_j1_{numero_partido}", 0))
        set1_j2 = int(form.get(f"set1_j2_{numero_partido}", 0))
        tb1_j1 = int(form.get(f"tb1_j1_{numero_partido}", 0))
        tb1_j2 = int(form.get(f"tb1_j2_{numero_partido}", 0))

        set2_j1 = int(form.get(f"set2_j1_{numero_partido}", 0))
        set2_j2 = int(form.get(f"set2_j2_{numero_partido}", 0))
        tb2_j1 = int(form.get(f"tb2_j1_{numero_partido}", 0))
        tb2_j2 = int(form.get(f"tb2_j2_{numero_partido}", 0))

        tipo_decisivo = form.get(f"tipo_decisivo_{numero_partido}", "")
        decisivo_j1 = int(form.get(f"decisivo_j1_{numero_partido}", 0))
        decisivo_j2 = int(form.get(f"decisivo_j2_{numero_partido}", 0))


        ganador1 = ganador_set(set1_j1, set1_j2)
        ganador2 = ganador_set(set2_j1, set2_j2)

    

        if ganador1 == 0 or ganador2 == 0:
            continue


        if (set1_j1, set1_j2) in [(7, 6), (6, 7)]:
            if not tiebreak_valido(tb1_j1, tb1_j2, 7):
                continue

        if (set2_j1, set2_j2) in [(7, 6), (6, 7)]:
            if not tiebreak_valido(tb2_j1, tb2_j2, 7):
                continue

        sets_j1 = 0
        sets_j2 = 0

        if ganador1 == 1:
            sets_j1 += 1
        else:
            sets_j2 += 1

        if ganador2 == 1:
            sets_j1 += 1
        else:
            sets_j2 += 1

        sets = [
            (1, set1_j1, set1_j2, tb1_j1, tb1_j2, 1),
            (2, set2_j1, set2_j2, tb2_j1, tb2_j2, 1)
        ]

        if sets_j1 == 1 and sets_j2 == 1:
            if tipo_decisivo == "super":
                if not tiebreak_valido(decisivo_j1, decisivo_j2, 10):
                    continue

                if decisivo_j1 > decisivo_j2:
                    ganador_id = jugador1_id
                    sets_j1 += 1
                else:
                    ganador_id = jugador2_id
                    sets_j2 += 1

                sets.append((3, decisivo_j1, decisivo_j2, 0, 0, 3))

            elif tipo_decisivo == "normal":
                ganador3 = ganador_set(decisivo_j1, decisivo_j2)

                if ganador3 == 0:
                    continue

                if ganador3 == 1:
                    ganador_id = jugador1_id
                    sets_j1 += 1
                else:
                    ganador_id = jugador2_id
                    sets_j2 += 1

                sets.append((3, decisivo_j1, decisivo_j2, 0, 0, 1))

            else:
                continue

        else:
                ganador_id = jugador1_id if sets_j1 == 2 else jugador2_id

        partes_resultado = []

        for numero_set, j1, j2, tbj1, tbj2, tipo_set in sets:
            if tipo_set == 3:
                partes_resultado.append(f"{j1}-{j2}")
            elif (j1, j2) in [(7, 6), (6, 7)]:
                partes_resultado.append(f"{j1}-{j2}({tbj1}-{tbj2})")
            else:
                partes_resultado.append(f"{j1}-{j2}")

        resultado = " ".join(partes_resultado)
        
        ronda_cuadro_id = guardar_o_actualizar_ronda_cuadro(
            cur,
            cuadro_id,
            1,
            "nombre_primera_ronda",
            numero_partido,
            jugador1_id,
            jugador2_id,
            ganador_id,
            jugador1_pos,
            jugador2_pos,
            "jugado",
            resultado,
            None
        )


        cur.execute("""
            SELECT id
            FROM partidos
            WHERE cuadro_id = %s
              AND ronda_numero = 1
              AND posicion_ronda = %s
        """, (cuadro_id, numero_partido))

        partido_existente = cur.fetchone()

        if partido_existente:
            partido_id = partido_existente[0]

            cur.execute("""
                UPDATE partidos
                SET ganador_id = %s,
                    resultado = %s,
                    estado = 'jugado',
                    ronda_cuadro_id = %s
                WHERE id = %s
            """, (ganador_id, resultado, ronda_cuadro_id, partido_id))

            cur.execute("""
                DELETE FROM sets
                WHERE partido_id = %s
            """, (partido_id,))
        else:
            cur.execute("""
                INSERT INTO partidos
                (
                    torneo_id,
                    fecha_partido,
                    jugador1_id,
                    jugador2_id,
                    ganador_id,
                    ronda,
                    resultado,
                    cuadro_id,
                    ronda_numero,
                    posicion_ronda,
                    jugador1_posicion,
                    jugador2_posicion,
                    estado,
                    ronda_cuadro_id
                )
                VALUES (%s, CURRENT_DATE, %s, %s, %s, %s, %s, %s, 1, %s, %s, %s, 'jugado', %s)
                RETURNING id
            """, (
                torneo_id,
                jugador1_id,
                jugador2_id,
                ganador_id,
                "nombre_primera_ronda",
                resultado,
                cuadro_id,
                numero_partido,
                jugador1_pos,
                jugador2_pos,
                ronda_cuadro_id
            ))

            partido_id = cur.fetchone()[0]

        cur.execute("""
            UPDATE rondas_cuadro
            SET partido_id = %s
            WHERE id = %s
        """, (partido_id, ronda_cuadro_id))

        for numero_set, j1, j2, tbj1, tbj2, tipo_set in sets:
            cur.execute("""
                INSERT INTO sets
                (
                    partido_id,
                    numero_set,
                    juegos_jugador1,
                    juegos_jugador2,
                    tiebreak_jugador1,
                    tiebreak_jugador2,
                    tipo_set
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                partido_id,
                numero_set,
                j1,
                j2,
                tbj1,
                tbj2,
                tipo_set
            ))


    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(
        url=f"/admin/cuadros/{cuadro_id}/resultados",
        status_code=303
    )
@app.post("/admin/cuadros/{cuadro_id}/guardar-resultados-ronda/{ronda_numero}")
async def guardar_resultados_ronda(
    cuadro_id: int,
    ronda_numero: int,
    request: Request,
    admin: str = Depends(comprobar_admin)
):
    form = await request.form()

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT torneo_id, tamano
        FROM cuadros
        WHERE id = %s
    """, (cuadro_id,))
    fila_cuadro = cur.fetchone()

    if not fila_cuadro:
        cur.close()
        conn.close()
        return RedirectResponse(url="/admin/cuadros", status_code=303)

    torneo_id = fila_cuadro[0]
    tamano_cuadro = fila_cuadro[1]

    nombres_rondas = nombres_rondas_por_tamano(tamano_cuadro)
    nombre_primera_ronda = nombres_rondas[0] if nombres_rondas else "Primera ronda"


    if ronda_numero <= len(nombres_rondas):
        nombre_ronda = nombres_rondas[ronda_numero - 1]
    else:
        cur.close()
        conn.close()
        return RedirectResponse(
            url=f"/admin/cuadros/{cuadro_id}/resultados",
            status_code=303
        )

    if ronda_numero == 1:
        prefijo = ""
    else:
        prefijo = f"r{ronda_numero}_"

    
    numeros_partido = []

    for key in form.keys():
        patron = f"jugador1_id_{prefijo}"
        if key.startswith(patron):
            numeros_partido.append(int(key.replace(patron, "")))

    for numero_partido in numeros_partido:
        jugador1_raw = form.get(f"jugador1_id_{prefijo}{numero_partido}", "")
        jugador2_raw = form.get(f"jugador2_id_{prefijo}{numero_partido}", "")

        # BYE jugador 1 pasa automáticamente
        if jugador1_raw and not jugador2_raw:

            jugador1_id = int(jugador1_raw)

            guardar_o_actualizar_bye(
                cur,
                torneo_id,
                cuadro_id,
                nombre_ronda,
                numero_partido,
                jugador1_id,
                None,
                jugador1_id,
                None,
                None
            )

            continue

        # BYE jugador 2 pasa automáticamente
        if jugador2_raw and not jugador1_raw:

            jugador2_id = int(jugador2_raw)

            guardar_o_actualizar_bye(
                cur,
                torneo_id,
                cuadro_id,
                nombre_ronda,
                numero_partido,
                None,
                jugador2_id,
                jugador2_id,
                None,
                None
            )

            continue

        # Si faltan los dos, no hacemos nada
        if not jugador1_raw and not jugador2_raw:
            continue

        jugador1_id = int(jugador1_raw)
        jugador2_id = int(jugador2_raw)




        tipo_resultado = form.get(
            f"tipo_resultado_{prefijo}{numero_partido}",
            "normal"
        )

        sets = []

        if tipo_resultado == "wo_j1":
            ganador_id = jugador2_id
            resultado = "WO J1"
            estado = "wo_jugador1"

        elif tipo_resultado == "wo_j2":
            ganador_id = jugador1_id
            resultado = "WO J2"
            estado = "wo_jugador2"

        else:
            set1_j1 = int(form.get(f"set1_j1_{prefijo}{numero_partido}", 0))
            set1_j2 = int(form.get(f"set1_j2_{prefijo}{numero_partido}", 0))
            tb1_j1 = int(form.get(f"tb1_j1_{prefijo}{numero_partido}", 0))
            tb1_j2 = int(form.get(f"tb1_j2_{prefijo}{numero_partido}", 0))

            set2_j1 = int(form.get(f"set2_j1_{prefijo}{numero_partido}", 0))
            set2_j2 = int(form.get(f"set2_j2_{prefijo}{numero_partido}", 0))
            tb2_j1 = int(form.get(f"tb2_j1_{prefijo}{numero_partido}", 0))
            tb2_j2 = int(form.get(f"tb2_j2_{prefijo}{numero_partido}", 0))

            tipo_decisivo = form.get(f"tipo_decisivo_{prefijo}{numero_partido}", "")
            decisivo_j1 = int(form.get(f"decisivo_j1_{prefijo}{numero_partido}", 0))
            decisivo_j2 = int(form.get(f"decisivo_j2_{prefijo}{numero_partido}", 0))

            if tipo_resultado == "ret_j1":
                ganador_id = jugador2_id
                resultado = f"{set1_j1}-{set1_j2} {set2_j1}-{set2_j2} RET J1"
                estado = "ret_jugador1"
                sets = [
                    (1, set1_j1, set1_j2, tb1_j1, tb1_j2, 1),
                    (2, set2_j1, set2_j2, tb2_j1, tb2_j2, 1)
                ]

            elif tipo_resultado == "ret_j2":
                ganador_id = jugador1_id
                resultado = f"{set1_j1}-{set1_j2} {set2_j1}-{set2_j2} RET J2"
                estado = "ret_jugador2"
                sets = [
                    (1, set1_j1, set1_j2, tb1_j1, tb1_j2, 1),
                    (2, set2_j1, set2_j2, tb2_j1, tb2_j2, 1)
                ]

            else:
                ganador1 = ganador_set(set1_j1, set1_j2)
                ganador2 = ganador_set(set2_j1, set2_j2)

                if ganador1 == 0 or ganador2 == 0:
                    continue

                sets_j1 = 1 if ganador1 == 1 else 0
                sets_j2 = 1 if ganador1 == 2 else 0

                if ganador2 == 1:
                    sets_j1 += 1
                else:
                    sets_j2 += 1

                sets = [
                    (1, set1_j1, set1_j2, tb1_j1, tb1_j2, 1),
                    (2, set2_j1, set2_j2, tb2_j1, tb2_j2, 1)
                ]

                if sets_j1 == 1 and sets_j2 == 1:
                    if tipo_decisivo == "super":
                        ganador_id = jugador1_id if decisivo_j1 > decisivo_j2 else jugador2_id
                        sets.append((3, decisivo_j1, decisivo_j2, 0, 0, 3))
                    else:
                        continue
                else:
                    ganador_id = jugador1_id if sets_j1 == 2 else jugador2_id

                partes = []
                for _, j1, j2, tbj1, tbj2, tipo_set in sets:
                    if (j1, j2) in [(7, 6), (6, 7)]:
                        partes.append(f"{j1}-{j2}({tbj1}-{tbj2})")
                    else:
                        partes.append(f"{j1}-{j2}")

                resultado = " ".join(partes)
                estado = "jugado"

        ronda_cuadro_id = guardar_o_actualizar_ronda_cuadro(
            cur,
            cuadro_id,
            ronda_numero,
            nombre_ronda,
            numero_partido,
            jugador1_id,
            jugador2_id,
            ganador_id,
            None,
            None,
            estado,
            resultado,
            None
        )

        cur.execute("""
            SELECT id
            FROM partidos
            WHERE cuadro_id = %s
              AND ronda_numero = %s
              AND posicion_ronda = %s
        """, (cuadro_id, ronda_numero, numero_partido))

        partido_existente = cur.fetchone()

        if partido_existente:
            partido_id = partido_existente[0]

            cur.execute("""
                UPDATE partidos
                SET jugador1_id = %s,
                    jugador2_id = %s,
                    ganador_id = %s,
                    resultado = %s,
                    ronda = %s,
                    estado = %s,
                    ronda_cuadro_id = %s
                WHERE id = %s
            """, (
                jugador1_id,
                jugador2_id,
                ganador_id,
                resultado,
                nombre_ronda,
                estado,
                ronda_cuadro_id,
                partido_id
            ))

            cur.execute("DELETE FROM sets WHERE partido_id = %s", (partido_id,))

        else:
            cur.execute("""
                INSERT INTO partidos
                (
                    torneo_id, fecha_partido, jugador1_id, jugador2_id,
                    ganador_id, ronda, resultado, cuadro_id,
                    ronda_numero, posicion_ronda, estado, ronda_cuadro_id
                )
                VALUES (%s, CURRENT_DATE, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                torneo_id,
                jugador1_id,
                jugador2_id,
                ganador_id,
                nombre_ronda,
                resultado,
                cuadro_id,
                ronda_numero,
                numero_partido,
                estado,
                ronda_cuadro_id
            ))

            partido_id = cur.fetchone()[0]

        cur.execute("""
            UPDATE rondas_cuadro
            SET partido_id = %s
            WHERE id = %s
        """, (partido_id, ronda_cuadro_id))

        for numero_set, j1, j2, tbj1, tbj2, tipo_set in sets:
            if j1 == 0 and j2 == 0:
                continue

            cur.execute("""
                INSERT INTO sets
                (
                    partido_id, numero_set, juegos_jugador1, juegos_jugador2,
                    tiebreak_jugador1, tiebreak_jugador2, tipo_set
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                partido_id,
                numero_set,
                j1,
                j2,
                tbj1,
                tbj2,
                tipo_set
            ))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(
        url=f"/admin/cuadros/{cuadro_id}/resultados",
        status_code=303
    )

@app.post("/admin/cuadros/{cuadro_id}/generar-siguiente-ronda")
def generar_siguiente_ronda_cuadro(
    cuadro_id: int,
    ronda_actual: int = Form(...),
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT tamano
        FROM cuadros
        WHERE id = %s
    """, (cuadro_id,))

    fila_tamano = cur.fetchone()
    tamano_cuadro = fila_tamano[0]

    nombres_rondas = nombres_rondas_por_tamano(tamano_cuadro)

    if ronda_actual < len(nombres_rondas):
        nombre_siguiente = nombres_rondas[ronda_actual]
    else:
        nombre_siguiente = "Final"


    generar_siguiente_ronda(
        cur,
        cuadro_id,
        ronda_actual,
        nombre_siguiente
    )

    conn.commit()

    cur.close()
    conn.close()

    return RedirectResponse(
        url=f"/admin/cuadros/{cuadro_id}/resultados",
        status_code=303
    )

@app.post("/admin/cuadros/{cuadro_id}/guardar-posiciones")
async def guardar_posiciones_cuadro(
    cuadro_id: int,
    request: Request,
    admin: str = Depends(comprobar_admin)
):
    form = await request.form()

    print("FORM RECIBIDO:", dict(form))

    conn = get_connection()
    cur = conn.cursor()

    for key, value in form.items():
        if key.startswith("posicion_"):
            inscrito_id = int(key.replace("posicion_", ""))

            if value:
                cur.execute("""
                    UPDATE cuadro_inscritos
                    SET posicion = %s
                    WHERE id = %s AND cuadro_id = %s
                """, (int(value), inscrito_id, cuadro_id))
            else:
                cur.execute("""
                    UPDATE cuadro_inscritos
                    SET posicion = NULL
                    WHERE id = %s AND cuadro_id = %s
                """, (inscrito_id, cuadro_id))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(
        url=f"/admin/cuadros/{cuadro_id}/inscritos",
        status_code=303
    )



@app.post("/admin/inscritos/{inscrito_id}/posicion")
def guardar_posicion_inscrito(
    inscrito_id: int,
    posicion: int = Form(...),
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT cuadro_id
        FROM cuadro_inscritos
        WHERE id = %s
    """, (inscrito_id,))
    fila = cur.fetchone()

    if not fila:
        cur.close()
        conn.close()
        return RedirectResponse(url="/admin/cuadros", status_code=303)

    cuadro_id = fila[0]

    cur.execute("""
        UPDATE cuadro_inscritos
        SET posicion = %s
        WHERE id = %s
    """, (posicion, inscrito_id))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(
        url=f"/admin/cuadros/{cuadro_id}/inscritos",
        status_code=303
    )

@app.post("/admin/cuadros/{cuadro_id}/importar-inscritos")
def importar_inscritos_desde_excel(
    cuadro_id: int,
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    # Buscar ruta_excel del cuadro
    cur.execute("""
        SELECT ruta_excel
        FROM cuadros
        WHERE id = %s
    """, (cuadro_id,))
    fila = cur.fetchone()

    if not fila or not fila[0]:
        cur.close()
        conn.close()
        return RedirectResponse(
            url=f"/admin/cuadros/{cuadro_id}/inscritos",
            status_code=303
        )

    ruta_excel = fila[0]

    # Si la ruta es relativa, la convertimos a absoluta desde la raíz del proyecto
    if not os.path.isabs(ruta_excel):
        ruta_excel = os.path.join(os.getcwd(), ruta_excel)

    wb = load_workbook(ruta_excel)
    ws = wb.active

    # Limpiar inscritos anteriores de ese cuadro para no duplicar
    cur.execute("""
        DELETE FROM cuadro_inscritos
        WHERE cuadro_id = %s
    """, (cuadro_id,))

    for fila_excel in ws.iter_rows(min_row=2):
        licencia = fila_excel[1].value  # Columna B
        nombre_excel = fila_excel[2].value  # Columna C

        if not licencia:
            continue

        licencia = str(licencia).strip().replace(".0", "")

        cur.execute("""
            SELECT id
            FROM jugadores
            WHERE TRIM(numero_licencia) = %s
        """, (licencia,))

        jugador = cur.fetchone()

        if jugador:
            jugador_id = jugador[0]
            estado = "encontrado"
        else:
            jugador_id = None
            estado = "no_encontrado"

        cur.execute("""
            INSERT INTO cuadro_inscritos
            (cuadro_id, jugador_id, numero_licencia, nombre_excel, estado)
            VALUES (%s, %s, %s, %s, %s)
        """, (cuadro_id, jugador_id, licencia, nombre_excel, estado))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(
        url=f"/admin/cuadros/{cuadro_id}/inscritos",
        status_code=303
    )

@app.post("/admin/cuadros")
def guardar_cuadro(
    torneo_id: int = Form(...),
    nombre: str = Form(...),
    categoria_id: int = Form(...),
    genero_id: int = Form(...),
    tamano: int = Form(...),
    numero_jugadores: int = Form(...),
    observaciones: str = Form(""),
    ruta_excel: str = Form(""),
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
    INSERT INTO cuadros (torneo_id, nombre, categoria_id, genero_id, tamano, numero_jugadores, observaciones, ruta_excel)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
""", (torneo_id, nombre, categoria_id, genero_id, tamano, numero_jugadores, observaciones, ruta_excel))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(url="/admin/cuadros", status_code=303)

@app.get("/admin/cuadros/editar/{cuadro_id}", response_class=HTMLResponse)
def editar_cuadro_form(
    request: Request,
    cuadro_id: int,
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, nombre, fecha_inicio, categoria, ubicacion
        FROM torneos
        ORDER BY fecha_inicio DESC, nombre
    """)
    torneos = cur.fetchall()

    cur.execute("""
        SELECT id, torneo_id, nombre, tamano, numero_jugadores, COALESCE(observaciones, ''), COALESCE(ruta_excel, ''), categoria_id, genero_id
        FROM cuadros
        WHERE id = %s
    """, (cuadro_id,))

    cuadro = cur.fetchone()

    cur.execute("""
        SELECT id, nombre
        FROM categorias
        ORDER BY id
    """)
    categorias = cur.fetchall()

    cur.execute("""
        SELECT id, nombre
        FROM generos
        ORDER BY id
    """)
    generos = cur.fetchall()

    cur.close()
    conn.close()

    if not cuadro:
        return RedirectResponse(url="/admin/cuadros", status_code=303)

    return templates.TemplateResponse(
        request=request,
        name="editar_cuadro.html",
        context={
            "request": request,
            "cuadro": cuadro,
            "torneos": torneos,
            "categorias": categorias,
            "generos": generos
        }
    )

@app.post("/admin/cuadros/editar/{cuadro_id}")
def actualizar_cuadro(
    cuadro_id: int,
    torneo_id: int = Form(...),
    nombre: str = Form(...),
    categoria_id: int = Form(...),
    genero_id: int = Form(...),
    tamano: int = Form(...),
    numero_jugadores: int = Form(...),
    observaciones: str = Form(""),
    ruta_excel: str = Form(""),
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE cuadros
        SET torneo_id = %s,
            nombre = %s,
            categoria_id = %s,
            genero_id = %s,
            tamano = %s,
            numero_jugadores = %s,
            observaciones = %s,
            ruta_excel = %s
        WHERE id = %s
    """, (torneo_id, nombre, categoria_id, genero_id, tamano, numero_jugadores, observaciones, ruta_excel, cuadro_id))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(url="/admin/cuadros", status_code=303)


@app.get("/admin/partidos", response_class=HTMLResponse)
def ver_partidos(
    request: Request,
    buscar: str = "",
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    texto_busqueda = f"%{buscar.strip()}%"

    cur.execute("""
        SELECT id, nombre, fecha_inicio, categoria, ubicacion
        FROM torneos
        ORDER BY id
    """)
    torneos = cur.fetchall()

    cur.execute("""
        SELECT id, nombre, apellido1, apellido2
        FROM jugadores
        ORDER BY apellido1, apellido2, nombre
    """)
    jugadores = cur.fetchall()

    cur.execute("""
        SELECT
            p.id,
            t.nombre AS torneo,
            COALESCE(c.nombre, '') AS cuadro,
            COALESCE(cat.nombre, '') AS categoria,
            COALESCE(gen.nombre, '') AS genero,
            j1.nombre || ' ' || j1.apellido1 || ' ' || COALESCE(j1.apellido2, '') AS jugador1,
            j2.nombre || ' ' || j2.apellido1 || ' ' || COALESCE(j2.apellido2, '') AS jugador2,
            COALESCE(g.nombre || ' ' || g.apellido1 || ' ' || COALESCE(g.apellido2, ''),'') AS ganador,
            CASE
                WHEN p.ronda ILIKE 'Treintaidosavos' THEN '1/32'
                WHEN p.ronda ILIKE 'Dieciseisavos' THEN '1/16'
                WHEN p.ronda ILIKE 'Octavos' THEN '1/8'
                WHEN p.ronda ILIKE 'Cuartos' THEN '1/4'
                WHEN p.ronda ILIKE 'Semifinal' THEN 'SF'
                WHEN p.ronda ILIKE 'Final' THEN 'F'
                ELSE p.ronda
            END AS ronda_corta,
            p.fecha_partido,
            p.resultado
        FROM partidos p
        JOIN torneos t ON p.torneo_id = t.id
        LEFT JOIN cuadros c ON p.cuadro_id = c.id
        LEFT JOIN categorias cat ON c.categoria_id = cat.id
        LEFT JOIN generos gen ON c.genero_id = gen.id
        JOIN jugadores j1 ON p.jugador1_id = j1.id
        JOIN jugadores j2 ON p.jugador2_id = j2.id
        LEFT JOIN jugadores g ON p.ganador_id = g.id
        WHERE
            %s = '%%'
            OR t.nombre ILIKE %s
            OR COALESCE(c.nombre, '') ILIKE %s
            OR COALESCE(cat.nombre, '') ILIKE %s
            OR COALESCE(gen.nombre, '') ILIKE %s
            OR j1.nombre ILIKE %s
            OR j1.apellido1 ILIKE %s
            OR COALESCE(j1.apellido2, '') ILIKE %s
            OR j2.nombre ILIKE %s
            OR j2.apellido1 ILIKE %s
            OR COALESCE(j2.apellido2, '') ILIKE %s
            OR COALESCE(g.nombre, '') ILIKE %s
            OR COALESCE(g.apellido1, '') ILIKE %s
            OR COALESCE(g.apellido2, '') ILIKE %s
            OR COALESCE(p.ronda, '') ILIKE %s
            OR COALESCE(p.resultado, '') ILIKE %s
        ORDER BY p.fecha_partido DESC, t.nombre, c.nombre, p.ronda_numero, p.posicion_ronda
    """, (
        texto_busqueda,
        texto_busqueda,
        texto_busqueda,
        texto_busqueda,
        texto_busqueda,
        texto_busqueda,
        texto_busqueda,
        texto_busqueda,
        texto_busqueda,
        texto_busqueda,
        texto_busqueda,
        texto_busqueda,
        texto_busqueda,
        texto_busqueda,
        texto_busqueda,
        texto_busqueda
    ))
                
    """)
    partidos = cur.fetchall()

    cur.close()
    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="partidos.html",
        context={
        
            "request": request,
            "torneos": torneos,
            "jugadores": jugadores,
            "partidos": partidos
            "buscar": buscar
        }
    )

@app.post("/admin/partidos")
def guardar_partido(
    torneo_id: int = Form(...),
    fecha_partido: str = Form(...),
    jugador1_id: int = Form(...),
    jugador2_id: int = Form(...),
    ganador_id: int = Form(...),
    ronda: str = Form(...),
    resultado: str = Form(...),
    admin: str = Depends(comprobar_admin)
):
    if jugador1_id == jugador2_id:
        return RedirectResponse(url="/admin/partidos", status_code=303)
    if ganador_id not in (jugador1_id, jugador2_id):
        return RedirectResponse(url="/admin/partidos", status_code=303)

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO partidos (
            torneo_id,
            fecha_partido,
            jugador1_id,
            jugador2_id,
            ganador_id,
            ronda,
            resultado
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (
        torneo_id,
        fecha_partido,
        jugador1_id,
        jugador2_id,
        ganador_id,
        ronda,
        resultado
    ))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(url="/admin/partidos", status_code=303)

@app.get("/admin/partidos/editar/{partido_id}", response_class=HTMLResponse)
def editar_partido_form(request: Request, partido_id: int,
admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, nombre, fecha_inicio, categoria, ubicacion
        FROM torneos
        ORDER BY id
    """)
    torneos = cur.fetchall()

    cur.execute("""
        SELECT id, nombre, apellido1, apellido2
        FROM jugadores
        ORDER BY apellido1, apellido2, nombre
    """)
    jugadores = cur.fetchall()

    cur.execute("""
        SELECT id, torneo_id, fecha_partido, jugador1_id, jugador2_id, ganador_id, ronda, resultado
        FROM partidos
        WHERE id = %s
    """, (partido_id,))
    partido = cur.fetchone()

    cur.close()
    conn.close()

    if not partido:
        return RedirectResponse(url="/admin/partidos", status_code=303)

    return templates.TemplateResponse(
        request=request,
        name="editar_partido.html",
        context={
            "request": request,
            "partido": partido,
            "torneos": torneos,
            "jugadores": jugadores
        }
    )


@app.post("/admin/partidos/editar/{partido_id}")
def actualizar_partido(
    partido_id: int,
    torneo_id: int = Form(...),
    fecha_partido: str = Form(...),
    jugador1_id: int = Form(...),
    jugador2_id: int = Form(...),
    ganador_id: int = Form(...),
    ronda: str = Form(...),
    resultado: str = Form(...),
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE partidos
        SET torneo_id = %s,
            fecha_partido = %s,
            jugador1_id = %s,
            jugador2_id = %s,
            ganador_id = %s,
            ronda = %s,
            resultado = %s
        WHERE id = %s
    """, (torneo_id, fecha_partido, jugador1_id, jugador2_id, ganador_id, ronda, resultado, partido_id))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(url="/admin/partidos", status_code=303)


@app.post("/admin/partidos/borrar/{partido_id}")
def borrar_partido(partido_id: int,
admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("DELETE FROM partidos WHERE id = %s", (partido_id,))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(url="/admin/partidos", status_code=303)

@app.get("/admin/sets", response_class=HTMLResponse)
def ver_sets(request: Request,
admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    # Traer sets ya preparados para mostrar bonito
    cur.execute("""
        SELECT
            s.id,
            s.partido_id,
            s.numero_set,
            s.juegos_jugador1,
            s.juegos_jugador2,
            COALESCE(s.tiebreak_jugador1, 0) AS tiebreak_jugador1,
            COALESCE(s.tiebreak_jugador2, 0) AS tiebreak_jugador2,
            s.tipo_set,

            p.jugador1_id,
            p.jugador2_id,
            p.ganador_id,
            p.ronda,

            j1.nombre || ' ' || j1.apellido1 || ' ' || COALESCE(j1.apellido2, '') AS jugador1,
            j2.nombre || ' ' || j2.apellido1 || ' ' || COALESCE(j2.apellido2, '') AS jugador2,
            g.nombre || ' ' || g.apellido1 || ' ' || COALESCE(g.apellido2, '') AS ganador
        FROM sets s
        JOIN partidos p ON s.partido_id = p.id
        JOIN jugadores j1 ON p.jugador1_id = j1.id
        JOIN jugadores j2 ON p.jugador2_id = j2.id
        JOIN jugadores g ON p.ganador_id = g.id
        ORDER BY s.partido_id, s.numero_set
    """)
    filas_sets = cur.fetchall()

    sets_bonitos = []

    for fila in filas_sets:
        (
            set_id,
            partido_id,
            numero_set,
            juegos_jugador1,
            juegos_jugador2,
            tiebreak_jugador1,
            tiebreak_jugador2,
            tipo_set,
            jugador1_id,
            jugador2_id,
            ganador_id,
            ronda,
            jugador1,
            jugador2,
            ganador
        ) = fila

        if ganador_id == jugador1_id:
            juegos_ganador = juegos_jugador1
            juegos_rival = juegos_jugador2
            tiebreak_ganador = tiebreak_jugador1
            tiebreak_rival = tiebreak_jugador2
        else:
            juegos_ganador = juegos_jugador2
            juegos_rival = juegos_jugador1
            tiebreak_ganador = tiebreak_jugador2
            tiebreak_rival = tiebreak_jugador1

        if tipo_set == 1:
            tipo_set_texto = "Set normal"
        elif tipo_set == 2:
            tipo_set_texto = "Set con tiebreak"
        elif tipo_set == 3:
            tipo_set_texto = "Super tiebreak"
        else:
            tipo_set_texto = "No definido"

        sets_bonitos.append({
            "id": set_id,
            "partido_id": partido_id,
            "partido": f"{jugador1.strip()} vs {jugador2.strip()}",
            "ganador": ganador.strip(),
            "ronda": ronda,
            "numero_set": numero_set,
            "juegos_ganador": juegos_ganador,
            "juegos_rival": juegos_rival,
            "resultado_set": f"{juegos_ganador}-{juegos_rival}",
            "tiebreak_ganador": tiebreak_ganador,
            "tiebreak_rival": tiebreak_rival,
            "tipo_set": tipo_set_texto
        })

    # Traer partidos para el desplegable del formulario
    cur.execute("""
        SELECT
            p.id,
            j1.nombre,
            j1.apellido1,
            j2.nombre,
            j2.apellido1,
            p.resultado,
            p.ronda,
            p.fecha_partido,
            g.nombre,
            g.apellido1
        FROM partidos p
        JOIN jugadores j1 ON p.jugador1_id = j1.id
        JOIN jugadores j2 ON p.jugador2_id = j2.id
        JOIN jugadores g ON p.ganador_id = g.id
        ORDER BY p.fecha_partido, p.id
    """)
    partidos = cur.fetchall()

    cur.close()
    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="sets.html",
        context={
            "request": request,
            "sets": sets_bonitos,
            "partidos": partidos
        }
    )

@app.post("/admin/sets")
def guardar_set(
    partido_id: int = Form(...),
    numero_set: int = Form(...),
    juegos_jugador1: int = Form(...),
    juegos_jugador2: int = Form(...),
    tiebreak_jugador1: int = Form(None),
    tiebreak_jugador2: int = Form(None),
    tipo_set: int = Form(...),
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()
    # Comprobar si ya existe ese set para ese partido
    cur.execute("""
        SELECT id
        FROM sets
        WHERE partido_id = %s AND numero_set = %s
    """, (partido_id, numero_set))

    set_existente = cur.fetchone()

    if set_existente:
        cur.close()
        conn.close()
        return RedirectResponse(url="/admin/sets", status_code=303)

    cur.execute("""
        INSERT INTO sets (
            partido_id,
            numero_set,
            juegos_jugador1,
            juegos_jugador2,
            tiebreak_jugador1,
            tiebreak_jugador2,
            tipo_set
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (
        partido_id,
        numero_set,
        juegos_jugador1,
        juegos_jugador2,
        tiebreak_jugador1,
        tiebreak_jugador2,
        tipo_set
    ))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(url="/admin/sets", status_code=303)

@app.get("/admin/sets/editar/{set_id}", response_class=HTMLResponse)
def editar_set_form(request: Request, set_id: int,
admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            p.id,
            j1.nombre || ' ' || j1.apellido1 || ' ' || COALESCE(j1.apellido2, '') AS jugador1,
            j2.nombre || ' ' || j2.apellido1 || ' ' || COALESCE(j2.apellido2, '') AS jugador2,
            p.resultado,
            p.ronda,
            p.fecha_partido
        FROM partidos p
        JOIN jugadores j1 ON p.jugador1_id = j1.id
        JOIN jugadores j2 ON p.jugador2_id = j2.id
        ORDER BY p.fecha_partido, p.id
    """)
    partidos = cur.fetchall()

    cur.execute("""
        SELECT
            id,
            partido_id,
            numero_set,
            juegos_jugador1,
            juegos_jugador2,
            COALESCE(tiebreak_jugador1, 0),
            COALESCE(tiebreak_jugador2, 0),
            tipo_set
        FROM sets
        WHERE id = %s
    """, (set_id,))
    set_item = cur.fetchone()

    cur.close()
    conn.close()

    if not set_item:
        return RedirectResponse(url="/admin/sets", status_code=303)

    return templates.TemplateResponse(
        request=request,
        name="editar_set.html",
        context={
            "request": request,
            "set_item": set_item,
            "partidos": partidos
        }
    )


@app.post("/admin/sets/editar/{set_id}")
def actualizar_set(
    set_id: int,
    partido_id: int = Form(...),
    numero_set: int = Form(...),
    juegos_jugador1: int = Form(...),
    juegos_jugador2: int = Form(...),
    tiebreak_jugador1: int = Form(0),
    tiebreak_jugador2: int = Form(0),
    tipo_set: int = Form(...),
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE sets
        SET partido_id = %s,
            numero_set = %s,
            juegos_jugador1 = %s,
            juegos_jugador2 = %s,
            tiebreak_jugador1 = %s,
            tiebreak_jugador2 = %s,
            tipo_set = %s
        WHERE id = %s
    """, (
        partido_id,
        numero_set,
        juegos_jugador1,
        juegos_jugador2,
        tiebreak_jugador1,
        tiebreak_jugador2,
        tipo_set,
        set_id
    ))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(url="/admin/sets", status_code=303)


@app.post("/admin/sets/borrar/{set_id}")
def borrar_set(set_id: int,
admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("DELETE FROM sets WHERE id = %s", (set_id,))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(url="/admin/sets", status_code=303)