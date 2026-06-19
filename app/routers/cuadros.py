import os
import re
import unicodedata
from datetime import date

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse
from openpyxl import load_workbook

from app.core import comprobar_admin, templates
from app.database import get_connection
from app.services.importador_cuadro_federacion import importar_pdf_cuadro_federacion

router = APIRouter()

@router.get("/admin/cuadros", response_class=HTMLResponse)
def ver_cuadros(
    request: Request,
    torneo_id: int = 0,
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, nombre, fecha_inicio, categoria, ubicacion
        FROM torneos
        ORDER BY fecha_inicio DESC, nombre
    """)
    torneos = cur.fetchall()

    cur.execute("""
        SELECT id, nombre
        FROM generos
        ORDER BY id
    """)
    generos = cur.fetchall()

    cur.execute("""
        SELECT id, nombre
        FROM categorias
        ORDER BY id
    """)
    categorias = cur.fetchall()

    if torneo_id != 0:
        cur.execute("""
            SELECT
                c.id,
                c.nombre,
                c.tamano,
                c.numero_jugadores,
                COALESCE(c.observaciones, ''),
                t.nombre,
                t.fecha_inicio,
                COALESCE(cat.nombre, ''),
                COALESCE(g.nombre, ''),
                t.ubicacion,
                
                (
                    SELECT COUNT(*)
                    FROM rondas_cuadro rc
                    WHERE rc.cuadro_id = c.id
                        AND rc.ganador_id IS NOT NULL
                ) AS resultados_guardados,

                (
                    SELECT TRIM(
                        j.nombre || ' ' ||
                        j.apellido1 || ' ' ||
                        COALESCE(j.apellido2, '')
                    )
                
                    FROM rondas_cuadro rc
                    JOIN jugadores j ON j.id = rc.ganador_id
                    WHERE rc.cuadro_id = c.id
                    ORDER BY rc.ronda_numero DESC,
                            rc.posicion_ronda DESC
                    LIMIT 1

                ) AS ganador    

                    

            FROM cuadros c
            JOIN torneos t ON c.torneo_id = t.id
            LEFT JOIN categorias cat ON c.categoria_id = cat.id
            LEFT JOIN generos g ON c.genero_id = g.id
            WHERE c.torneo_id = %s
            ORDER BY cat.nombre, g.nombre, c.nombre
        """, (torneo_id,))
    else:
        cur.execute("""
            SELECT
                c.id,
                c.nombre,
                c.tamano,
                c.numero_jugadores,
                COALESCE(c.observaciones, ''),
                t.nombre,
                t.fecha_inicio,
                COALESCE(cat.nombre, ''),
                COALESCE(g.nombre, ''),
                t.ubicacion,

                (
                    SELECT COUNT(*)
                    FROM rondas_cuadro rc
                    WHERE rc.cuadro_id = c.id
                        AND rc.ganador_id IS NOT NULL
                ) AS resultados_guardados,

                (
                    SELECT TRIM(
                        j.nombre || ' ' ||
                        j.apellido1 || ' ' ||
                        COALESCE(j.apellido2, '')
                    )
                
                    FROM rondas_cuadro rc
                    JOIN jugadores j ON j.id = rc.ganador_id
                    WHERE rc.cuadro_id = c.id
                    ORDER BY rc.ronda_numero DESC,
                            rc.posicion_ronda DESC
                    LIMIT 1

                ) AS ganador
                    
            FROM cuadros c
            JOIN torneos t ON c.torneo_id = t.id
            LEFT JOIN categorias cat ON c.categoria_id = cat.id
            LEFT JOIN generos g ON c.genero_id = g.id
            ORDER BY t.fecha_inicio DESC, t.nombre, cat.nombre, g.nombre, c.nombre
        """)   

    cuadros = cur.fetchall()

    cur.close()
    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="cuadros.html",
        context={
            "request": request,
            "torneos": torneos,
            "cuadros": cuadros,
            "torneo_id": torneo_id,
            "generos": generos,
            "categorias": categorias
        }
    )

@router.get("/admin/cuadros/{cuadro_id}/importar-federacion-prueba", response_class=HTMLResponse)
def probar_importacion_federacion_form(
    cuadro_id: int,
    admin: str = Depends(comprobar_admin)
):
    return HTMLResponse(f"""
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <title>Probar importación Federación</title>
        </head>
        <body>
            <h1>Probar importación Federación Catalana</h1>
            <p>Esta prueba solo extrae texto y muestra una pantalla de revision. No guarda ni modifica datos.</p>
            <form action="/admin/cuadros/{cuadro_id}/importar-federacion-prueba" method="post" enctype="multipart/form-data">
                <label>PDF del cuadro:</label>
                <input type="file" name="file" accept=".pdf" required>
                <br><br>
                <label>XLSX de inscritos:</label>
                <input type="file" name="inscritos_file" accept=".xlsx" required>
                <br><br>
                <button type="submit">Analizar PDF y XLSX</button>
            </form>
            <br>
            <a href="/admin/cuadros/{cuadro_id}/inscritos">Volver a inscritos</a>
        </body>
        </html>
    """)


@router.post("/admin/cuadros/{cuadro_id}/importar-federacion-prueba")
async def probar_importacion_federacion_pdf(
    request: Request,
    cuadro_id: int,
    file: UploadFile = File(...),
    inscritos_file: UploadFile = File(...),
    admin: str = Depends(comprobar_admin)
):
    contenido = await file.read()
    resultado = importar_pdf_cuadro_federacion(contenido)
    inscritos_revision = leer_inscritos_excel_para_revision(inscritos_file)
    partidos = cruzar_partidos_con_inscritos(
        resultado.get("partidos_ronda_1", []),
        inscritos_revision,
    )
    destino_importacion = preparar_destino_importacion(
        resultado,
        inscritos_revision,
        inscritos_file.filename,
    )
    resultado["cuadro_id"] = cuadro_id
    resultado["archivo"] = file.filename
    resultado["archivo_inscritos"] = inscritos_file.filename
    resultado["modo"] = "solo_prueba_sin_guardar"
    return templates.TemplateResponse(
        request=request,
        name="revision_importacion_federacion.html",
        context={
            "request": request,
            "cuadro_id": cuadro_id,
            "resultado": resultado,
            "entradas": resultado.get("ronda_1", []),
            "partidos": partidos,
            "inscritos": inscritos_revision,
            "destino_importacion": destino_importacion,
        },
    )


def preparar_destino_importacion(
    resultado: dict,
    inscritos_revision: dict[int, dict],
    nombre_archivo_excel: str | None,
) -> dict:
    cabecera = resultado.get("cabecera", {})
    entradas = resultado.get("ronda_1", [])
    texto_cabecera = " ".join(cabecera.get("lineas_cabecera", []))
    modalidad = cabecera.get("modalidad_categoria_genero_cuadro") or ""
    texto_deteccion = " ".join([texto_cabecera, modalidad])

    torneo_nombre = cabecera.get("torneo") or ""
    fecha_inicio = extraer_fecha_inicio(cabecera.get("fechas", []))
    ubicacion = cabecera.get("club") or ""
    categoria_nombre = detectar_categoria_pdf(texto_deteccion)
    genero_nombre = detectar_genero_pdf(texto_deteccion)
    cuadro_nombre = detectar_nombre_cuadro_pdf(texto_deteccion)
    tamano = detectar_tamano_cuadro(entradas)
    numero_jugadores = sum(1 for entrada in entradas if not entrada.get("bye"))

    conn = get_connection()
    cur = conn.cursor()

    torneo_existente = buscar_torneo_revision(
        cur,
        torneo_nombre,
        fecha_inicio,
        ubicacion,
    )
    categoria = buscar_catalogo_revision(cur, "categorias", categoria_nombre)
    genero = buscar_catalogo_revision(cur, "generos", genero_nombre)

    cuadro_existente = None
    if torneo_existente and categoria and genero and cuadro_nombre:
        cur.execute("""
            SELECT id
            FROM cuadros
            WHERE torneo_id = %s
              AND LOWER(TRIM(nombre)) = LOWER(TRIM(%s))
              AND categoria_id = %s
              AND genero_id = %s
            LIMIT 1
        """, (torneo_existente["id"], cuadro_nombre, categoria["id"], genero["id"]))
        fila_cuadro = cur.fetchone()
        if fila_cuadro:
            cuadro_existente = {"id": fila_cuadro[0]}

    cur.close()
    conn.close()

    errores = []
    if not categoria:
        errores.append("Categoria detectada no encontrada en la base de datos.")
    if not genero:
        errores.append("Genero detectado no encontrado en la base de datos.")

    return {
        "torneo": {
            "nombre": torneo_nombre,
            "fecha_inicio": fecha_inicio,
            "ubicacion": ubicacion,
            "existe": bool(torneo_existente),
            "torneo_id": torneo_existente["id"] if torneo_existente else None,
            "accion": "usar_existente" if torneo_existente else "crear_nuevo",
        },
        "categoria": {
            "detectada": categoria_nombre,
            "categoria_id": categoria["id"] if categoria else None,
            "nombre_bd": categoria["nombre"] if categoria else None,
            "error": not bool(categoria),
        },
        "genero": {
            "detectado": genero_nombre,
            "genero_id": genero["id"] if genero else None,
            "nombre_bd": genero["nombre"] if genero else None,
            "error": not bool(genero),
        },
        "cuadro": {
            "nombre": cuadro_nombre,
            "tamano": tamano,
            "numero_jugadores": numero_jugadores,
            "ruta_excel": nombre_archivo_excel or "",
            "observaciones": modalidad,
            "existe": bool(cuadro_existente),
            "cuadro_id": cuadro_existente["id"] if cuadro_existente else None,
            "accion": "usar_existente" if cuadro_existente else "crear_nuevo",
        },
        "inscritos_detectados": len(inscritos_revision),
        "errores": errores,
        "puede_confirmar": not errores,
    }


def buscar_torneo_revision(
    cur,
    nombre: str,
    fecha_inicio: date | None,
    ubicacion: str,
) -> dict | None:
    if not nombre:
        return None

    if fecha_inicio:
        cur.execute("""
            SELECT id
            FROM torneos
            WHERE LOWER(TRIM(nombre)) = LOWER(TRIM(%s))
              AND fecha_inicio = %s
              AND LOWER(TRIM(COALESCE(ubicacion, ''))) = LOWER(TRIM(%s))
            LIMIT 1
        """, (nombre, fecha_inicio, ubicacion or ""))
    else:
        cur.execute("""
            SELECT id
            FROM torneos
            WHERE LOWER(TRIM(nombre)) = LOWER(TRIM(%s))
              AND LOWER(TRIM(COALESCE(ubicacion, ''))) = LOWER(TRIM(%s))
            LIMIT 1
        """, (nombre, ubicacion or ""))

    fila = cur.fetchone()
    if not fila:
        return None

    return {"id": fila[0]}


def buscar_catalogo_revision(cur, tabla: str, nombre: str | None) -> dict | None:
    if tabla not in {"categorias", "generos"} or not nombre:
        return None

    cur.execute(f"""
        SELECT id, nombre
        FROM {tabla}
    """)

    nombre_normalizado = normalizar_nombre_revision(nombre)
    for fila in cur.fetchall():
        if normalizar_nombre_revision(fila[1]) == nombre_normalizado:
            return {"id": fila[0], "nombre": fila[1]}

    return None


def extraer_fecha_inicio(fechas: list[str]) -> date | None:
    if not fechas:
        return None

    fecha = fechas[0].strip()
    match_numerica = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", fecha)
    if match_numerica:
        dia = int(match_numerica.group(1))
        mes = int(match_numerica.group(2))
        ano = int(match_numerica.group(3))
        if ano < 100:
            ano += 2000
        return date(ano, mes, dia)

    meses = {
        "GENER": 1,
        "ENERO": 1,
        "FEBRER": 2,
        "FEBRERO": 2,
        "MARC": 3,
        "MARZO": 3,
        "ABRIL": 4,
        "MAIG": 5,
        "MAYO": 5,
        "JUNY": 6,
        "JUNIO": 6,
        "JULIOL": 7,
        "JULIO": 7,
        "AGOST": 8,
        "AGOSTO": 8,
        "SETEMBRE": 9,
        "SEPTIEMBRE": 9,
        "OCTUBRE": 10,
        "OCTUBRE": 10,
        "NOVEMBRE": 11,
        "NOVIEMBRE": 11,
        "DESEMBRE": 12,
        "DICIEMBRE": 12,
    }
    fecha_normalizada = normalizar_nombre_revision(fecha)
    match_textual = re.search(r"(\d{1,2}) DE ([A-Z]+)(?: DE)? (\d{4})", fecha_normalizada)
    if match_textual and match_textual.group(2) in meses:
        return date(
            int(match_textual.group(3)),
            meses[match_textual.group(2)],
            int(match_textual.group(1)),
        )

    return None


def detectar_categoria_pdf(texto: str) -> str | None:
    normalizado = normalizar_nombre_revision(texto)
    if "SUB 10" in normalizado or "SUB10" in normalizado:
        return "benjamin"
    if "SUB 12" in normalizado or "SUB12" in normalizado:
        return "alevin"
    if "SUB 14" in normalizado or "SUB14" in normalizado:
        return "infantil"
    return None


def detectar_genero_pdf(texto: str) -> str | None:
    normalizado = normalizar_nombre_revision(texto)
    if "MASCULINO" in normalizado or "MASCULI" in normalizado:
        return "masculino"
    if "FEMENINO" in normalizado or "FEMENI" in normalizado:
        return "femenino"
    return None


def detectar_nombre_cuadro_pdf(texto: str) -> str:
    normalizado = normalizar_nombre_revision(texto)
    match = re.search(r"\b(?:CUADRO|QUADRE)\s*(\d+)\b", normalizado)
    if match:
        return f"Cuadro {match.group(1)}"
    return "Cuadro 1"


def detectar_tamano_cuadro(entradas: list[dict]) -> int:
    posiciones = [
        entrada.get("posicion")
        for entrada in entradas
        if entrada.get("posicion") is not None
    ]
    if not posiciones:
        return 0

    max_posicion = max(posiciones)
    for tamano in (8, 16, 32, 64, 128):
        if max_posicion <= tamano:
            return tamano

    return max_posicion


def leer_inscritos_excel_para_revision(file: UploadFile) -> dict[int, dict]:
    conn = get_connection()
    cur = conn.cursor()
    wb = load_workbook(file.file, data_only=True)
    ws = wb.active
    inscritos = {}
    posicion = 1

    for fila_excel in ws.iter_rows(min_row=2):
        licencia = fila_excel[1].value  # Columna B
        nombre_excel = fila_excel[2].value  # Columna C

        if not licencia:
            continue

        licencia = str(licencia).strip().replace(".0", "")
        nombre_excel = str(nombre_excel).strip() if nombre_excel else ""

        cur.execute("""
            SELECT
                id,
                nombre,
                apellido1,
                apellido2,
                numero_licencia
            FROM jugadores
            WHERE TRIM(numero_licencia) = %s
        """, (licencia,))

        jugador = cur.fetchone()
        jugador_id = None
        nombre_oficial = None
        licencia_oficial = None
        estado = "no_encontrado"

        if jugador:
            jugador_id = jugador[0]
            nombre_oficial = nombre_completo_jugador(jugador)
            licencia_oficial = jugador[4]
            estado = "encontrado"

        inscritos[posicion] = {
            "posicion": posicion,
            "numero_licencia": licencia,
            "nombre_excel": nombre_excel,
            "jugador_id": jugador_id,
            "nombre_oficial": nombre_oficial,
            "numero_licencia_oficial": licencia_oficial,
            "estado": estado,
        }
        posicion += 1

    cur.close()
    conn.close()
    return inscritos


def cruzar_partidos_con_inscritos(
    partidos: list[dict],
    inscritos_por_posicion: dict[int, dict],
) -> list[dict]:
    partidos_cruzados = []
    inscritos_usados = set()

    for partido in partidos:
        partido_cruzado = dict(partido)
        inscrito1 = buscar_inscrito_para_pdf(
            partido.get("jugador1_detectado"),
            partido.get("posicion_jugador1"),
            inscritos_por_posicion,
            inscritos_usados,
        )
        inscrito2 = buscar_inscrito_para_pdf(
            partido.get("jugador2_detectado"),
            partido.get("posicion_jugador2"),
            inscritos_por_posicion,
            inscritos_usados,
        )

        partido_cruzado.update(datos_inscrito_para_lado(
            "jugador1",
            partido.get("jugador1_detectado"),
            inscrito1,
        ))
        partido_cruzado.update(datos_inscrito_para_lado(
            "jugador2",
            partido.get("jugador2_detectado"),
            inscrito2,
        ))
        partidos_cruzados.append(partido_cruzado)

    return partidos_cruzados


def buscar_inscrito_para_pdf(
    nombre_pdf: str | None,
    posicion_fallback: int | None,
    inscritos_por_posicion: dict[int, dict],
    inscritos_usados: set,
) -> dict | None:
    if not nombre_pdf or nombre_pdf.upper().startswith("BYE"):
        return None

    mejor_inscrito = None
    mejor_puntuacion = 0

    for inscrito in inscritos_por_posicion.values():
        clave = inscrito.get("numero_licencia") or inscrito.get("posicion")
        if clave in inscritos_usados:
            continue

        puntuacion = max(
            puntuacion_nombre_revision(nombre_pdf, inscrito.get("nombre_oficial")),
            puntuacion_nombre_revision(nombre_pdf, inscrito.get("nombre_excel")),
        )

        if puntuacion > mejor_puntuacion:
            mejor_puntuacion = puntuacion
            mejor_inscrito = inscrito

    if mejor_inscrito and mejor_puntuacion >= 2:
        clave = mejor_inscrito.get("numero_licencia") or mejor_inscrito.get("posicion")
        inscritos_usados.add(clave)
        return mejor_inscrito

    inscrito_fallback = inscritos_por_posicion.get(posicion_fallback)
    if inscrito_fallback:
        clave = inscrito_fallback.get("numero_licencia") or inscrito_fallback.get("posicion")
        if clave not in inscritos_usados:
            inscritos_usados.add(clave)
            return inscrito_fallback

    return None


def datos_inscrito_para_lado(
    prefijo: str,
    nombre_pdf: str | None,
    inscrito: dict | None,
) -> dict:
    nombre_oficial = inscrito.get("nombre_oficial") if inscrito else None
    nombre_excel = inscrito.get("nombre_excel") if inscrito else None
    numero_licencia = inscrito.get("numero_licencia") if inscrito else None
    jugador_id = inscrito.get("jugador_id") if inscrito else None
    estado = inscrito.get("estado") if inscrito else "sin_inscrito"

    return {
        f"{prefijo}_numero_licencia": numero_licencia,
        f"{prefijo}_nombre_excel": nombre_excel,
        f"{prefijo}_jugador_id_oficial": jugador_id,
        f"{prefijo}_oficial": nombre_oficial,
        f"{prefijo}_estado_inscrito": estado,
        f"{prefijo}_comparacion_nombre": comparar_nombres_revision(
            nombre_pdf,
            nombre_oficial,
        ),
    }


def nombre_completo_jugador(jugador: tuple) -> str:
    return " ".join(
        parte
        for parte in (jugador[1], jugador[2], jugador[3])
        if parte
    ).strip()


def comparar_nombres_revision(nombre_pdf: str | None, nombre_oficial: str | None) -> str:
    if not nombre_pdf:
        return "sin_pdf"

    if not nombre_oficial:
        return "sin_oficial"

    pdf = normalizar_nombre_revision(nombre_pdf)
    oficial = normalizar_nombre_revision(nombre_oficial)

    if pdf == oficial:
        return "coincide"

    partes_pdf = set(pdf.split())
    partes_oficial = set(oficial.split())
    partes_comunes = partes_pdf.intersection(partes_oficial)

    if len(partes_comunes) >= 2:
        return "probable"

    return "revisar"


def puntuacion_nombre_revision(nombre_pdf: str | None, nombre_candidato: str | None) -> int:
    if not nombre_pdf or not nombre_candidato:
        return 0

    partes_pdf = set(normalizar_nombre_revision(nombre_pdf).split())
    partes_candidato = set(normalizar_nombre_revision(nombre_candidato).split())
    return len(partes_pdf.intersection(partes_candidato))


def normalizar_nombre_revision(nombre: str) -> str:
    nombre = unicodedata.normalize("NFKD", nombre)
    nombre = "".join(char for char in nombre if not unicodedata.combining(char))
    nombre = re.sub(r"[^A-Z0-9 ]+", " ", nombre.upper())
    return re.sub(r"\s+", " ", nombre).strip()


@router.post("/admin/cuadros/{cuadro_id}/importar-excel-archivo")
def importar_excel_archivo(
    cuadro_id: int,
    file: UploadFile = File(...),
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    wb = load_workbook(file.file)
    ws = wb.active

    # Borrar inscritos anteriores del cuadro para no duplicar
    cur.execute("""
        DELETE FROM cuadro_inscritos
        WHERE cuadro_id = %s
    """, (cuadro_id,))

    for fila_excel in ws.iter_rows(min_row=2):
        licencia = fila_excel[1].value  # columna B
        nombre_excel = fila_excel[2].value  # columna C

        if not licencia:
            continue

        licencia = str(licencia).strip().replace(".0", "")

        cur.execute("""
            SELECT id
            FROM jugadores
            WHERE TRIM(numero_licencia) = %s
        """, (licencia,))

        jugador = cur.fetchone()

        if jugador:
            jugador_id = jugador[0]
            estado = "encontrado"
        else:
            jugador_id = None
            estado = "no_encontrado"

        cur.execute("""
            INSERT INTO cuadro_inscritos
            (cuadro_id, jugador_id, numero_licencia, nombre_excel, estado)
            VALUES (%s, %s, %s, %s, %s)
        """, (cuadro_id, jugador_id, licencia, nombre_excel, estado))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(
        url=f"/admin/cuadros/{cuadro_id}/inscritos",
        status_code=303
    )
@router.get("/admin/cuadros/{cuadro_id}/inscritos", response_class=HTMLResponse)
def ver_inscritos(
    cuadro_id: int,
    request: Request,
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            ci.id,
            ci.numero_licencia,
            ci.nombre_excel,
            j.id,
            j.nombre,
            j.apellido1,
            j.apellido2,
            ci.estado,
            ci.posicion
        FROM cuadro_inscritos ci
        LEFT JOIN jugadores j ON ci.jugador_id = j.id
        WHERE ci.cuadro_id = %s
        ORDER BY ci.posicion NULLS LAST, ci.nombre_excel
    """, (cuadro_id,))

    inscritos = cur.fetchall()

    cur.execute("""
        SELECT tamano
        FROM cuadros
        WHERE id = %s
    """, (cuadro_id,))

    cuadro = cur.fetchone()
    tamano_cuadro = cuadro[0] if cuadro else 32
    posiciones = list(range(1, tamano_cuadro + 1))

    posiciones_ocupadas = [
        i[8] for i in inscritos
        if i[8] is not None
    ]

    inscritos_por_posicion = {
        i[8]: i
        for i in inscritos
        if i[8] is not None
    }

    cuadro_ordenado = []

    for posicion in posiciones:
        inscrito = inscritos_por_posicion.get(posicion)

        if inscrito:
            cuadro_ordenado.append({
                "posicion": posicion,
                "tipo": "jugador",
                "inscrito": inscrito
            })
        else:
            cuadro_ordenado.append({
                "posicion": posicion,
                "tipo": "bye",
                "inscrito": None
            })

    cur.close()
    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="inscritos.html",
        context={
            "request": request,
            "inscritos": inscritos,
            "cuadro_id": cuadro_id,
            "posiciones": posiciones,
            "posiciones_ocupadas": posiciones_ocupadas,
            "cuadro_ordenado": cuadro_ordenado
        }
    )

@router.get("/admin/cuadros/{cuadro_id}/resultados", response_class=HTMLResponse)
def resultados_cuadro(
    cuadro_id: int,
    request: Request,
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            ci.id,
            ci.numero_licencia,
            ci.nombre_excel,
            j.id,
            j.nombre,
            j.apellido1,
            j.apellido2,
            ci.estado,
            ci.posicion
        FROM cuadro_inscritos ci
        LEFT JOIN jugadores j ON ci.jugador_id = j.id
        WHERE ci.cuadro_id = %s
        ORDER BY ci.posicion NULLS LAST, ci.nombre_excel
    """, (cuadro_id,))

    inscritos = cur.fetchall()

    cur.execute("""
        SELECT tamano
        FROM cuadros
        WHERE id = %s
    """, (cuadro_id,))

    cuadro = cur.fetchone()
    tamano_cuadro = cuadro[0] if cuadro else 32
    posiciones = list(range(1, tamano_cuadro + 1))

    inscritos_por_posicion = {
        i[8]: i
        for i in inscritos
        if i[8] is not None
    }

    cuadro_ordenado = []

    for posicion in posiciones:
        inscrito = inscritos_por_posicion.get(posicion)

        if inscrito:
            cuadro_ordenado.append({
                "posicion": posicion,
                "tipo": "jugador",
                "inscrito": inscrito
            })
        else:
            cuadro_ordenado.append({
                "posicion": posicion,
                "tipo": "bye",
                "inscrito": None
            })

    cur.execute("""
        SELECT
            p.posicion_ronda,
            p.estado,
            p.resultado,
            p.id
        FROM partidos p
        WHERE p.cuadro_id = %s
          AND p.ronda_numero = 1
    """, (cuadro_id,))

    partidos_guardados = {
        fila[0]: {
            "estado": fila[1],
            "resultado": fila[2],
            "partido_id": fila[3]
        }
        for fila in cur.fetchall()
    }

    emparejamientos = []
    numero_partido = 1

    for i in range(0, len(cuadro_ordenado), 2):
        lado1 = cuadro_ordenado[i]
        lado2 = cuadro_ordenado[i + 1] if i + 1 < len(cuadro_ordenado) else {
            "posicion": None,
            "tipo": "bye",
            "inscrito": None
        }

        if lado1["tipo"] == "jugador" and lado2["tipo"] == "jugador":
            estado = "partido"
        elif lado1["tipo"] == "jugador" and lado2["tipo"] == "bye":
            estado = "bye_jugador1"
        elif lado1["tipo"] == "bye" and lado2["tipo"] == "jugador":
            estado = "bye_jugador2"
        else:
            estado = "vacio"

        guardado = partidos_guardados.get(numero_partido)

        emparejamientos.append({
            "numero_partido": numero_partido,
            "lado1": lado1,
            "lado2": lado2,
            "estado": estado,
            "guardado": guardado
        })

        numero_partido += 1

    partidos_ronda_1 = []

    for partido in emparejamientos:
        jugador1_id = None
        jugador2_id = None
        jugador1_nombre = "BYE"
        jugador2_nombre = "BYE"


        if partido["lado1"]["tipo"] == "jugador":
            inscrito1 = partido["lado1"]["inscrito"]

            jugador1_id = inscrito1[3]

            jugador1_nombre = (
                f"{inscrito1[4]} "
                f"{inscrito1[5]} "
                f"{inscrito1[6] or ''}"
            ).strip()

        if partido["lado2"]["tipo"] == "jugador":
            inscrito2 = partido["lado2"]["inscrito"]

            jugador2_id = inscrito2[3]

            jugador2_nombre = (
                f"{inscrito2[4]} "
                f"{inscrito2[5]} "
                f"{inscrito2[6] or ''}"
            ).strip()


        partidos_ronda_1.append({
            "numero_partido": partido["numero_partido"],
            "jugador1_id": jugador1_id,
            "jugador1_nombre": jugador1_nombre,
            "jugador2_id": jugador2_id,
            "jugador2_nombre": jugador2_nombre,
            "ganador_id": None,
            "resultado": partido["guardado"]["resultado"] if partido["guardado"] else None,
            "estado": partido["estado"]
        })

    nombres_rondas = nombres_rondas_por_tamano(tamano_cuadro)
    rondas = []

    for numero_ronda, nombre_ronda in enumerate(nombres_rondas, start=1):
        prefijo = "" if numero_ronda == 1 else f"r{numero_ronda}_"

        if numero_ronda == 1:
            rondas.append({
                "numero": 1,
                "nombre": nombre_ronda,
                "prefijo": prefijo,
                "es_final": len(nombres_rondas) == 1,
                "partidos": partidos_ronda_1
            })
            continue

        cur.execute("""
            SELECT
                rc.posicion_ronda,
                rc.jugador1_id,
                TRIM(j1.nombre || ' ' || j1.apellido1 || ' ' || COALESCE(j1.apellido2, '')) AS jugador1_nombre,
                rc.jugador2_id,
                TRIM(j2.nombre || ' ' || j2.apellido1 || ' ' || COALESCE(j2.apellido2, '')) AS jugador2_nombre,
                rc.ganador_id,
                rc.resultado,
                rc.estado
            FROM rondas_cuadro rc
            LEFT JOIN jugadores j1 ON rc.jugador1_id = j1.id
            LEFT JOIN jugadores j2 ON rc.jugador2_id = j2.id
            WHERE rc.cuadro_id = %s
              AND rc.ronda_numero = %s
            ORDER BY rc.posicion_ronda
        """, (cuadro_id, numero_ronda))

        partidos = []

        for fila in cur.fetchall():
            partidos.append({
                "numero_partido": fila[0],
                "jugador1_id": fila[1],
                "jugador1_nombre": fila[2],
                "jugador2_id": fila[3],
                "jugador2_nombre": fila[4],
                "ganador_id": fila[5],
                "resultado": fila[6],
                "estado": fila[7]
            })

        if partidos:
            rondas.append({
                "numero": numero_ronda,
                "nombre": nombre_ronda,
                "prefijo": prefijo,
                "es_final": numero_ronda == len(nombres_rondas),
                "partidos": partidos
            })

    cur.close()
    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="resultados_cuadro.html",
        context={
            "request": request,
            "cuadro_id": cuadro_id,
            "emparejamientos": emparejamientos,
            "rondas": rondas
        }
    )

def guardar_o_actualizar_ronda_cuadro(
    cur,
    cuadro_id,
    ronda_numero,
    nombre_ronda,
    posicion_ronda,
    jugador1_id,
    jugador2_id,
    ganador_id,
    jugador1_posicion,
    jugador2_posicion,
    estado,
    resultado,
    partido_id=None
):
    cur.execute("""
        INSERT INTO rondas_cuadro
        (
            cuadro_id,
            ronda_numero,
            nombre_ronda,
            posicion_ronda,
            jugador1_id,
            jugador2_id,
            ganador_id,
            jugador1_posicion,
            jugador2_posicion,
            estado,
            resultado,
            partido_id
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (cuadro_id, ronda_numero, posicion_ronda)
        DO UPDATE SET
            nombre_ronda = EXCLUDED.nombre_ronda,
            jugador1_id = EXCLUDED.jugador1_id,
            jugador2_id = EXCLUDED.jugador2_id,
            ganador_id = EXCLUDED.ganador_id,
            jugador1_posicion = EXCLUDED.jugador1_posicion,
            jugador2_posicion = EXCLUDED.jugador2_posicion,
            estado = EXCLUDED.estado,
            resultado = EXCLUDED.resultado,
            partido_id = EXCLUDED.partido_id
        RETURNING id
    """, (
        cuadro_id,
        ronda_numero,
        nombre_ronda,
        posicion_ronda,
        jugador1_id,
        jugador2_id,
        ganador_id,
        jugador1_posicion,
        jugador2_posicion,
        estado,
        resultado,
        partido_id
    ))

    return cur.fetchone()[0]

def nombres_rondas_por_tamano(tamano):

    if tamano == 8:
        return [
            "Cuartos",
            "Semifinal",
            "Final"
        ]

    elif tamano == 16:
        return [
            "Octavos",
            "Cuartos",
            "Semifinal",
            "Final"
        ]

    elif tamano == 32:
        return [
            "Dieciseisavos",
            "Octavos",
            "Cuartos",
            "Semifinal",
            "Final"
        ]

    elif tamano == 64:
        return [
            "Treintaidosavos",
            "Dieciseisavos",
            "Octavos",
            "Cuartos",
            "Semifinal",
            "Final"
        ]

    elif tamano == 128:
        return [
            "Sesentaicuatroavos",
            "Treintaidosavos",
            "Dieciseisavos",
            "Octavos",
            "Cuartos",
            "Semifinal",
            "Final"
        ]

    return []

def generar_siguiente_ronda(
    cur,
    cuadro_id,
    ronda_actual,
    nombre_siguiente_ronda
):
    ronda_siguiente = ronda_actual + 1

    cur.execute("""
        SELECT
            rc.posicion_ronda,
            COALESCE(rc.ganador_id, p.ganador_id) AS ganador_id
        FROM rondas_cuadro rc
        LEFT JOIN partidos p ON p.id = rc.partido_id
        WHERE rc.cuadro_id = %s
          AND rc.ronda_numero = %s
        ORDER BY rc.posicion_ronda
    """, (cuadro_id, ronda_actual))

    partidos_ronda = cur.fetchall()

    for i in range(0, len(partidos_ronda), 2):

        if i + 1 >= len(partidos_ronda):
            break

        ganador1 = partidos_ronda[i][1]
        ganador2 = partidos_ronda[i + 1][1]

        if not ganador1 or not ganador2:
            continue

        numero_partido = (i // 2) + 1

        cur.execute("""
            INSERT INTO rondas_cuadro
            (
                cuadro_id,
                ronda_numero,
                nombre_ronda,
                posicion_ronda,
                jugador1_id,
                jugador2_id,
                estado
            )
            VALUES (%s, %s, %s, %s, %s, %s, 'pendiente')
            ON CONFLICT (cuadro_id, ronda_numero, posicion_ronda)
            DO UPDATE SET
                jugador1_id = EXCLUDED.jugador1_id,
                jugador2_id = EXCLUDED.jugador2_id,
                estado = 'pendiente'
        """, (
            cuadro_id,
            ronda_siguiente,
            nombre_siguiente_ronda,
            numero_partido,
            ganador1,
            ganador2
        ))

def ganador_set(j1, j2):
    # 6-0, 6-1, 6-2, 6-3, 6-4
    if j1 == 6 and j2 <= 4:
        return 1

    if j2 == 6 and j1 <= 4:
        return 2

    # 7-5
    if j1 == 7 and j2 == 5:
        return 1

    if j2 == 7 and j1 == 5:
        return 2

    # 7-6 / 6-7
    if j1 == 7 and j2 == 6:
        return 1

    if j2 == 7 and j1 == 6:
        return 2

    return 0



def tiebreak_valido(j1, j2, minimo):
    ganador = max(j1, j2)
    perdedor = min(j1, j2)
    return ganador >= minimo and ganador - perdedor >= 2

def guardar_o_actualizar_bye(
    cur,
    torneo_id,
    cuadro_id,
    nombre_primera_ronda,
    numero_partido,
    jugador1_id,
    jugador2_id,
    ganador_id,
    jugador1_pos,
    jugador2_pos
):
    ronda_cuadro_id = guardar_o_actualizar_ronda_cuadro(
        cur,
        cuadro_id,
        1,
        nombre_primera_ronda,
        numero_partido,
        jugador1_id,
        jugador2_id,
        ganador_id,
        jugador1_pos,
        jugador2_pos,
        "bye",
        "BYE",
        None
    )

    cur.execute("""
        SELECT id
        FROM partidos
        WHERE cuadro_id = %s
          AND ronda_numero = 1
          AND posicion_ronda = %s
    """, (cuadro_id, numero_partido))

    partido_existente = cur.fetchone()

    if partido_existente:
        partido_id = partido_existente[0]

        cur.execute("""
            UPDATE partidos
            SET jugador1_id = %s,
                jugador2_id = %s,
                ganador_id = %s,
                ronda = %s,
                resultado = %s,
                jugador1_posicion = %s,
                jugador2_posicion = %s,
                estado = 'bye',
                ronda_cuadro_id = %s
            WHERE id = %s
        """, (
            jugador1_id,
            jugador2_id,
            ganador_id,
            nombre_primera_ronda,
            "BYE",
            jugador1_pos,
            jugador2_pos,
            ronda_cuadro_id,
            partido_id
        ))

        cur.execute("""
            DELETE FROM sets
            WHERE partido_id = %s
        """, (partido_id,))

    else:
        cur.execute("""
            INSERT INTO partidos
            (
                torneo_id,
                fecha_partido,
                jugador1_id,
                jugador2_id,
                ganador_id,
                ronda,
                resultado,
                cuadro_id,
                ronda_numero,
                posicion_ronda,
                jugador1_posicion,
                jugador2_posicion,
                estado,
                ronda_cuadro_id
            )
            VALUES (%s, CURRENT_DATE, %s, %s, %s, %s, %s, %s, 1, %s, %s, %s, 'bye', %s)
            RETURNING id
        """, (
            torneo_id,
            jugador1_id,
            jugador2_id,
            ganador_id,
            nombre_primera_ronda,
            "BYE",
            cuadro_id,
            numero_partido,
            jugador1_pos,
            jugador2_pos,
            ronda_cuadro_id
        ))
        partido_id = cur.fetchone()[0]

    cur.execute("""
        UPDATE rondas_cuadro
        SET partido_id = %s
        WHERE id = %s
    """, (partido_id, ronda_cuadro_id))    


@router.post("/admin/cuadros/{cuadro_id}/guardar-resultados")
async def guardar_resultados_cuadro(
    cuadro_id: int,
    request: Request,
    admin: str = Depends(comprobar_admin)
):
    form = await request.form()
    
    print("FORM RECIBIDO:", dict(form))

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT torneo_id
        FROM cuadros
        WHERE id = %s
    """, (cuadro_id,))
    fila_cuadro = cur.fetchone()

    if not fila_cuadro:
        cur.close()
        conn.close()
        return RedirectResponse(url="/admin/cuadros", status_code=303)
    


    torneo_id = fila_cuadro[0]

    cur.execute("""
        SELECT
            ci.id,
            j.id,
            ci.posicion
        FROM cuadro_inscritos ci
        LEFT JOIN jugadores j ON ci.jugador_id = j.id
        WHERE ci.cuadro_id = %s
          AND ci.posicion IS NOT NULL
        ORDER BY ci.posicion
    """, (cuadro_id,))

    inscritos_posiciones = cur.fetchall()

    jugadores_por_posicion = {
        fila[2]: fila[1]
        for fila in inscritos_posiciones
    }

    cur.execute("""
        SELECT tamano
        FROM cuadros
        WHERE id = %s
    """, (cuadro_id,))

    fila_tamano = cur.fetchone()
    tamano_cuadro = fila_tamano[0] if fila_tamano else 32
    nombres_rondas = nombres_rondas_por_tamano(tamano_cuadro)

    nombre_primera_ronda = nombres_rondas[0] if nombres_rondas else "Primera ronda"
    numero_partido_bye = 1

    for posicion in range(1, tamano_cuadro + 1, 2):
        pos1 = posicion
        pos2 = posicion + 1

        jugador1_id = jugadores_por_posicion.get(pos1)
        jugador2_id = jugadores_por_posicion.get(pos2)

        if jugador1_id and not jugador2_id:
            print("GUARDANDO BYE J1:", numero_partido_bye, pos1, pos2, jugador1_id)    
            guardar_o_actualizar_bye(
                cur,
                torneo_id,
                cuadro_id,
                nombre_primera_ronda,
                numero_partido_bye,
                jugador1_id,
                None,
                jugador1_id,
                pos1,
                pos2
            )

        elif jugador2_id and not jugador1_id:
            print("GUARDANDO BYE J2:", numero_partido_bye, pos1, pos2, jugador2_id)
            guardar_o_actualizar_bye(
                cur,
                torneo_id,
                cuadro_id,
                nombre_primera_ronda,
                numero_partido_bye,
                None,
                jugador2_id,
                jugador2_id,
                pos1,
                pos2
            )

        numero_partido_bye += 1

    numeros_partido = []

    for key in form.keys():
        if key.startswith("jugador1_id_") and not key.startswith("jugador1_id_r"):
            numeros_partido.append(int(key.replace("jugador1_id_", "")))

    for numero_partido in numeros_partido:
        jugador1_id = int(form.get(f"jugador1_id_{numero_partido}"))
        jugador2_id = int(form.get(f"jugador2_id_{numero_partido}"))
        jugador1_pos = int(form.get(f"jugador1_pos_{numero_partido}"))
        jugador2_pos = int(form.get(f"jugador2_pos_{numero_partido}"))

        set1_j1 = int(form.get(f"set1_j1_{numero_partido}", 0))
        set1_j2 = int(form.get(f"set1_j2_{numero_partido}", 0))
        tb1_j1 = int(form.get(f"tb1_j1_{numero_partido}", 0))
        tb1_j2 = int(form.get(f"tb1_j2_{numero_partido}", 0))

        set2_j1 = int(form.get(f"set2_j1_{numero_partido}", 0))
        set2_j2 = int(form.get(f"set2_j2_{numero_partido}", 0))
        tb2_j1 = int(form.get(f"tb2_j1_{numero_partido}", 0))
        tb2_j2 = int(form.get(f"tb2_j2_{numero_partido}", 0))

        tipo_decisivo = form.get(f"tipo_decisivo_{numero_partido}", "")
        decisivo_j1 = int(form.get(f"decisivo_j1_{numero_partido}", 0))
        decisivo_j2 = int(form.get(f"decisivo_j2_{numero_partido}", 0))


        ganador1 = ganador_set(set1_j1, set1_j2)
        ganador2 = ganador_set(set2_j1, set2_j2)

    

        if ganador1 == 0 or ganador2 == 0:
            continue


        if (set1_j1, set1_j2) in [(7, 6), (6, 7)]:
            if not tiebreak_valido(tb1_j1, tb1_j2, 7):
                continue

        if (set2_j1, set2_j2) in [(7, 6), (6, 7)]:
            if not tiebreak_valido(tb2_j1, tb2_j2, 7):
                continue

        sets_j1 = 0
        sets_j2 = 0

        if ganador1 == 1:
            sets_j1 += 1
        else:
            sets_j2 += 1

        if ganador2 == 1:
            sets_j1 += 1
        else:
            sets_j2 += 1

        sets = [
            (1, set1_j1, set1_j2, tb1_j1, tb1_j2, 1),
            (2, set2_j1, set2_j2, tb2_j1, tb2_j2, 1)
        ]

        if sets_j1 == 1 and sets_j2 == 1:
            if tipo_decisivo == "super":
                if not tiebreak_valido(decisivo_j1, decisivo_j2, 10):
                    continue

                if decisivo_j1 > decisivo_j2:
                    ganador_id = jugador1_id
                    sets_j1 += 1
                else:
                    ganador_id = jugador2_id
                    sets_j2 += 1

                sets.append((3, decisivo_j1, decisivo_j2, 0, 0, 3))

            elif tipo_decisivo == "normal":
                ganador3 = ganador_set(decisivo_j1, decisivo_j2)

                if ganador3 == 0:
                    continue

                if ganador3 == 1:
                    ganador_id = jugador1_id
                    sets_j1 += 1
                else:
                    ganador_id = jugador2_id
                    sets_j2 += 1

                sets.append((3, decisivo_j1, decisivo_j2, 0, 0, 1))

            else:
                continue

        else:
                ganador_id = jugador1_id if sets_j1 == 2 else jugador2_id

        partes_resultado = []

        for numero_set, j1, j2, tbj1, tbj2, tipo_set in sets:
            if tipo_set == 3:
                partes_resultado.append(f"{j1}-{j2}")
            elif (j1, j2) in [(7, 6), (6, 7)]:
                partes_resultado.append(f"{j1}-{j2}({tbj1}-{tbj2})")
            else:
                partes_resultado.append(f"{j1}-{j2}")

        resultado = " ".join(partes_resultado)
        
        ronda_cuadro_id = guardar_o_actualizar_ronda_cuadro(
            cur,
            cuadro_id,
            1,
            "nombre_primera_ronda",
            numero_partido,
            jugador1_id,
            jugador2_id,
            ganador_id,
            jugador1_pos,
            jugador2_pos,
            "jugado",
            resultado,
            None
        )


        cur.execute("""
            SELECT id
            FROM partidos
            WHERE cuadro_id = %s
              AND ronda_numero = 1
              AND posicion_ronda = %s
        """, (cuadro_id, numero_partido))

        partido_existente = cur.fetchone()

        if partido_existente:
            partido_id = partido_existente[0]

            cur.execute("""
                UPDATE partidos
                SET ganador_id = %s,
                    resultado = %s,
                    estado = 'jugado',
                    ronda_cuadro_id = %s
                WHERE id = %s
            """, (ganador_id, resultado, ronda_cuadro_id, partido_id))

            cur.execute("""
                DELETE FROM sets
                WHERE partido_id = %s
            """, (partido_id,))
        else:
            cur.execute("""
                INSERT INTO partidos
                (
                    torneo_id,
                    fecha_partido,
                    jugador1_id,
                    jugador2_id,
                    ganador_id,
                    ronda,
                    resultado,
                    cuadro_id,
                    ronda_numero,
                    posicion_ronda,
                    jugador1_posicion,
                    jugador2_posicion,
                    estado,
                    ronda_cuadro_id
                )
                VALUES (%s, CURRENT_DATE, %s, %s, %s, %s, %s, %s, 1, %s, %s, %s, 'jugado', %s)
                RETURNING id
            """, (
                torneo_id,
                jugador1_id,
                jugador2_id,
                ganador_id,
                "nombre_primera_ronda",
                resultado,
                cuadro_id,
                numero_partido,
                jugador1_pos,
                jugador2_pos,
                ronda_cuadro_id
            ))

            partido_id = cur.fetchone()[0]

        cur.execute("""
            UPDATE rondas_cuadro
            SET partido_id = %s
            WHERE id = %s
        """, (partido_id, ronda_cuadro_id))

        for numero_set, j1, j2, tbj1, tbj2, tipo_set in sets:
            cur.execute("""
                INSERT INTO sets
                (
                    partido_id,
                    numero_set,
                    juegos_jugador1,
                    juegos_jugador2,
                    tiebreak_jugador1,
                    tiebreak_jugador2,
                    tipo_set
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                partido_id,
                numero_set,
                j1,
                j2,
                tbj1,
                tbj2,
                tipo_set
            ))


    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(
        url=f"/admin/cuadros/{cuadro_id}/resultados",
        status_code=303
    )
@router.post("/admin/cuadros/{cuadro_id}/guardar-resultados-ronda/{ronda_numero}")
async def guardar_resultados_ronda(
    cuadro_id: int,
    ronda_numero: int,
    request: Request,
    admin: str = Depends(comprobar_admin)
):
    form = await request.form()

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT torneo_id, tamano
        FROM cuadros
        WHERE id = %s
    """, (cuadro_id,))
    fila_cuadro = cur.fetchone()

    if not fila_cuadro:
        cur.close()
        conn.close()
        return RedirectResponse(url="/admin/cuadros", status_code=303)

    torneo_id = fila_cuadro[0]
    tamano_cuadro = fila_cuadro[1]

    nombres_rondas = nombres_rondas_por_tamano(tamano_cuadro)
    nombre_primera_ronda = nombres_rondas[0] if nombres_rondas else "Primera ronda"


    if ronda_numero <= len(nombres_rondas):
        nombre_ronda = nombres_rondas[ronda_numero - 1]
    else:
        cur.close()
        conn.close()
        return RedirectResponse(
            url=f"/admin/cuadros/{cuadro_id}/resultados",
            status_code=303
        )

    if ronda_numero == 1:
        prefijo = ""
    else:
        prefijo = f"r{ronda_numero}_"

    
    numeros_partido = []

    for key in form.keys():
        patron = f"jugador1_id_{prefijo}"
        if key.startswith(patron):
            numeros_partido.append(int(key.replace(patron, "")))

    for numero_partido in numeros_partido:
        jugador1_raw = form.get(f"jugador1_id_{prefijo}{numero_partido}", "")
        jugador2_raw = form.get(f"jugador2_id_{prefijo}{numero_partido}", "")

        # BYE jugador 1 pasa automáticamente
        if jugador1_raw and not jugador2_raw:

            jugador1_id = int(jugador1_raw)

            guardar_o_actualizar_bye(
                cur,
                torneo_id,
                cuadro_id,
                nombre_ronda,
                numero_partido,
                jugador1_id,
                None,
                jugador1_id,
                None,
                None
            )

            continue

        # BYE jugador 2 pasa automáticamente
        if jugador2_raw and not jugador1_raw:

            jugador2_id = int(jugador2_raw)

            guardar_o_actualizar_bye(
                cur,
                torneo_id,
                cuadro_id,
                nombre_ronda,
                numero_partido,
                None,
                jugador2_id,
                jugador2_id,
                None,
                None
            )

            continue

        # Si faltan los dos, no hacemos nada
        if not jugador1_raw and not jugador2_raw:
            continue

        jugador1_id = int(jugador1_raw)
        jugador2_id = int(jugador2_raw)




        tipo_resultado = form.get(
            f"tipo_resultado_{prefijo}{numero_partido}",
            "normal"
        )

        sets = []

        if tipo_resultado == "wo_j1":
            ganador_id = jugador2_id
            resultado = "WO J1"
            estado = "wo_jugador1"

        elif tipo_resultado == "wo_j2":
            ganador_id = jugador1_id
            resultado = "WO J2"
            estado = "wo_jugador2"

        else:
            set1_j1 = int(form.get(f"set1_j1_{prefijo}{numero_partido}", 0))
            set1_j2 = int(form.get(f"set1_j2_{prefijo}{numero_partido}", 0))
            tb1_j1 = int(form.get(f"tb1_j1_{prefijo}{numero_partido}", 0))
            tb1_j2 = int(form.get(f"tb1_j2_{prefijo}{numero_partido}", 0))

            set2_j1 = int(form.get(f"set2_j1_{prefijo}{numero_partido}", 0))
            set2_j2 = int(form.get(f"set2_j2_{prefijo}{numero_partido}", 0))
            tb2_j1 = int(form.get(f"tb2_j1_{prefijo}{numero_partido}", 0))
            tb2_j2 = int(form.get(f"tb2_j2_{prefijo}{numero_partido}", 0))

            tipo_decisivo = form.get(f"tipo_decisivo_{prefijo}{numero_partido}", "")
            decisivo_j1 = int(form.get(f"decisivo_j1_{prefijo}{numero_partido}", 0))
            decisivo_j2 = int(form.get(f"decisivo_j2_{prefijo}{numero_partido}", 0))

            if tipo_resultado == "ret_j1":
                ganador_id = jugador2_id
                resultado = f"{set1_j1}-{set1_j2} {set2_j1}-{set2_j2} RET J1"
                estado = "ret_jugador1"
                sets = [
                    (1, set1_j1, set1_j2, tb1_j1, tb1_j2, 1),
                    (2, set2_j1, set2_j2, tb2_j1, tb2_j2, 1)
                ]

            elif tipo_resultado == "ret_j2":
                ganador_id = jugador1_id
                resultado = f"{set1_j1}-{set1_j2} {set2_j1}-{set2_j2} RET J2"
                estado = "ret_jugador2"
                sets = [
                    (1, set1_j1, set1_j2, tb1_j1, tb1_j2, 1),
                    (2, set2_j1, set2_j2, tb2_j1, tb2_j2, 1)
                ]

            else:
                ganador1 = ganador_set(set1_j1, set1_j2)
                ganador2 = ganador_set(set2_j1, set2_j2)

                if ganador1 == 0 or ganador2 == 0:
                    continue

                sets_j1 = 1 if ganador1 == 1 else 0
                sets_j2 = 1 if ganador1 == 2 else 0

                if ganador2 == 1:
                    sets_j1 += 1
                else:
                    sets_j2 += 1

                sets = [
                    (1, set1_j1, set1_j2, tb1_j1, tb1_j2, 1),
                    (2, set2_j1, set2_j2, tb2_j1, tb2_j2, 1)
                ]

                if sets_j1 == 1 and sets_j2 == 1:
                    if tipo_decisivo == "super":
                        if decisivo_j1 == decisivo_j2:
                            continue

                        ganador_id = jugador1_id if decisivo_j1 > decisivo_j2 else jugador2_id
                        sets.append((3, decisivo_j1, decisivo_j2, 0, 0, 3))

                    elif tipo_decisivo == "normal":
                        ganador3 = ganador_set(decisivo_j1, decisivo_j2)

                        if ganador3 == 0:
                            continue

                        ganador_id = jugador1_id if ganador3 == 1 else jugador2_id
                        sets.append((3, decisivo_j1, decisivo_j2, 0, 0, 1))

                    else:
                        continue
                else:
                    ganador_id = jugador1_id if sets_j1 == 2 else jugador2_id


                partes = []
                for _, j1, j2, tbj1, tbj2, tipo_set in sets:
                    if (j1, j2) in [(7, 6), (6, 7)]:
                        partes.append(f"{j1}-{j2}({tbj1}-{tbj2})")
                    else:
                        partes.append(f"{j1}-{j2}")

                resultado = " ".join(partes)
                estado = "jugado"

        ronda_cuadro_id = guardar_o_actualizar_ronda_cuadro(
            cur,
            cuadro_id,
            ronda_numero,
            nombre_ronda,
            numero_partido,
            jugador1_id,
            jugador2_id,
            ganador_id,
            None,
            None,
            estado,
            resultado,
            None
        )

        cur.execute("""
            SELECT id
            FROM partidos
            WHERE cuadro_id = %s
              AND ronda_numero = %s
              AND posicion_ronda = %s
        """, (cuadro_id, ronda_numero, numero_partido))

        partido_existente = cur.fetchone()

        if partido_existente:
            partido_id = partido_existente[0]

            cur.execute("""
                UPDATE partidos
                SET jugador1_id = %s,
                    jugador2_id = %s,
                    ganador_id = %s,
                    resultado = %s,
                    ronda = %s,
                    estado = %s,
                    ronda_cuadro_id = %s
                WHERE id = %s
            """, (
                jugador1_id,
                jugador2_id,
                ganador_id,
                resultado,
                nombre_ronda,
                estado,
                ronda_cuadro_id,
                partido_id
            ))

            cur.execute("DELETE FROM sets WHERE partido_id = %s", (partido_id,))

        else:
            cur.execute("""
                INSERT INTO partidos
                (
                    torneo_id, fecha_partido, jugador1_id, jugador2_id,
                    ganador_id, ronda, resultado, cuadro_id,
                    ronda_numero, posicion_ronda, estado, ronda_cuadro_id
                )
                VALUES (%s, CURRENT_DATE, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                torneo_id,
                jugador1_id,
                jugador2_id,
                ganador_id,
                nombre_ronda,
                resultado,
                cuadro_id,
                ronda_numero,
                numero_partido,
                estado,
                ronda_cuadro_id
            ))

            partido_id = cur.fetchone()[0]

        cur.execute("""
            UPDATE rondas_cuadro
            SET partido_id = %s
            WHERE id = %s
        """, (partido_id, ronda_cuadro_id))

        for numero_set, j1, j2, tbj1, tbj2, tipo_set in sets:
            if j1 == 0 and j2 == 0:
                continue

            cur.execute("""
                INSERT INTO sets
                (
                    partido_id, numero_set, juegos_jugador1, juegos_jugador2,
                    tiebreak_jugador1, tiebreak_jugador2, tipo_set
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                partido_id,
                numero_set,
                j1,
                j2,
                tbj1,
                tbj2,
                tipo_set
            ))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(
        url=f"/admin/cuadros/{cuadro_id}/resultados",
        status_code=303
    )

@router.post("/admin/cuadros/{cuadro_id}/generar-siguiente-ronda")
def generar_siguiente_ronda_cuadro(
    cuadro_id: int,
    ronda_actual: int = Form(...),
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT tamano
        FROM cuadros
        WHERE id = %s
    """, (cuadro_id,))

    fila_tamano = cur.fetchone()
    tamano_cuadro = fila_tamano[0]

    nombres_rondas = nombres_rondas_por_tamano(tamano_cuadro)

    if ronda_actual < len(nombres_rondas):
        nombre_siguiente = nombres_rondas[ronda_actual]
    else:
        nombre_siguiente = "Final"


    generar_siguiente_ronda(
        cur,
        cuadro_id,
        ronda_actual,
        nombre_siguiente
    )

    conn.commit()

    cur.close()
    conn.close()

    return RedirectResponse(
        url=f"/admin/cuadros/{cuadro_id}/resultados",
        status_code=303
    )

@router.post("/admin/cuadros/{cuadro_id}/guardar-posiciones")
async def guardar_posiciones_cuadro(
    cuadro_id: int,
    request: Request,
    admin: str = Depends(comprobar_admin)
):
    form = await request.form()

    print("FORM RECIBIDO:", dict(form))

    conn = get_connection()
    cur = conn.cursor()

    for key, value in form.items():
        if key.startswith("posicion_"):
            inscrito_id = int(key.replace("posicion_", ""))

            if value:
                cur.execute("""
                    UPDATE cuadro_inscritos
                    SET posicion = %s
                    WHERE id = %s AND cuadro_id = %s
                """, (int(value), inscrito_id, cuadro_id))
            else:
                cur.execute("""
                    UPDATE cuadro_inscritos
                    SET posicion = NULL
                    WHERE id = %s AND cuadro_id = %s
                """, (inscrito_id, cuadro_id))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(
        url=f"/admin/cuadros/{cuadro_id}/inscritos",
        status_code=303
    )



@router.post("/admin/inscritos/{inscrito_id}/posicion")
def guardar_posicion_inscrito(
    inscrito_id: int,
    posicion: int = Form(...),
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT cuadro_id
        FROM cuadro_inscritos
        WHERE id = %s
    """, (inscrito_id,))
    fila = cur.fetchone()

    if not fila:
        cur.close()
        conn.close()
        return RedirectResponse(url="/admin/cuadros", status_code=303)

    cuadro_id = fila[0]

    cur.execute("""
        UPDATE cuadro_inscritos
        SET posicion = %s
        WHERE id = %s
    """, (posicion, inscrito_id))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(
        url=f"/admin/cuadros/{cuadro_id}/inscritos",
        status_code=303
    )

@router.post("/admin/cuadros/{cuadro_id}/importar-inscritos")
def importar_inscritos_desde_excel(
    cuadro_id: int,
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    # Buscar ruta_excel del cuadro
    cur.execute("""
        SELECT ruta_excel
        FROM cuadros
        WHERE id = %s
    """, (cuadro_id,))
    fila = cur.fetchone()

    if not fila or not fila[0]:
        cur.close()
        conn.close()
        return RedirectResponse(
            url=f"/admin/cuadros/{cuadro_id}/inscritos",
            status_code=303
        )

    ruta_excel = fila[0]

    # Si la ruta es relativa, la convertimos a absoluta desde la raíz del proyecto
    if not os.path.isabs(ruta_excel):
        ruta_excel = os.path.join(os.getcwd(), ruta_excel)

    wb = load_workbook(ruta_excel)
    ws = wb.active

    # Limpiar inscritos anteriores de ese cuadro para no duplicar
    cur.execute("""
        DELETE FROM cuadro_inscritos
        WHERE cuadro_id = %s
    """, (cuadro_id,))

    for fila_excel in ws.iter_rows(min_row=2):
        licencia = fila_excel[1].value  # Columna B
        nombre_excel = fila_excel[2].value  # Columna C

        if not licencia:
            continue

        licencia = str(licencia).strip().replace(".0", "")

        cur.execute("""
            SELECT id
            FROM jugadores
            WHERE TRIM(numero_licencia) = %s
        """, (licencia,))

        jugador = cur.fetchone()

        if jugador:
            jugador_id = jugador[0]
            estado = "encontrado"
        else:
            jugador_id = None
            estado = "no_encontrado"

        cur.execute("""
            INSERT INTO cuadro_inscritos
            (cuadro_id, jugador_id, numero_licencia, nombre_excel, estado)
            VALUES (%s, %s, %s, %s, %s)
        """, (cuadro_id, jugador_id, licencia, nombre_excel, estado))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(
        url=f"/admin/cuadros/{cuadro_id}/inscritos",
        status_code=303
    )

@router.post("/admin/cuadros")
def guardar_cuadro(
    torneo_id: int = Form(...),
    nombre: str = Form(...),
    categoria_id: int = Form(...),
    genero_id: int = Form(...),
    tamano: int = Form(...),
    numero_jugadores: int = Form(...),
    observaciones: str = Form(""),
    ruta_excel: str = Form(""),
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
    INSERT INTO cuadros (torneo_id, nombre, categoria_id, genero_id, tamano, numero_jugadores, observaciones, ruta_excel)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
""", (torneo_id, nombre, categoria_id, genero_id, tamano, numero_jugadores, observaciones, ruta_excel))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(url="/admin/cuadros", status_code=303)

@router.get("/admin/cuadros/editar/{cuadro_id}", response_class=HTMLResponse)
def editar_cuadro_form(
    request: Request,
    cuadro_id: int,
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, nombre, fecha_inicio, categoria, ubicacion
        FROM torneos
        ORDER BY fecha_inicio DESC, nombre
    """)
    torneos = cur.fetchall()

    cur.execute("""
        SELECT id, torneo_id, nombre, tamano, numero_jugadores, COALESCE(observaciones, ''), COALESCE(ruta_excel, ''), categoria_id, genero_id
        FROM cuadros
        WHERE id = %s
    """, (cuadro_id,))

    cuadro = cur.fetchone()

    cur.execute("""
        SELECT id, nombre
        FROM categorias
        ORDER BY id
    """)
    categorias = cur.fetchall()

    cur.execute("""
        SELECT id, nombre
        FROM generos
        ORDER BY id
    """)
    generos = cur.fetchall()

    cur.close()
    conn.close()

    if not cuadro:
        return RedirectResponse(url="/admin/cuadros", status_code=303)

    return templates.TemplateResponse(
        request=request,
        name="editar_cuadro.html",
        context={
            "request": request,
            "cuadro": cuadro,
            "torneos": torneos,
            "categorias": categorias,
            "generos": generos
        }
    )

@router.post("/admin/cuadros/editar/{cuadro_id}")
def actualizar_cuadro(
    cuadro_id: int,
    torneo_id: int = Form(...),
    nombre: str = Form(...),
    categoria_id: int = Form(...),
    genero_id: int = Form(...),
    tamano: int = Form(...),
    numero_jugadores: int = Form(...),
    observaciones: str = Form(""),
    ruta_excel: str = Form(""),
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE cuadros
        SET torneo_id = %s,
            nombre = %s,
            categoria_id = %s,
            genero_id = %s,
            tamano = %s,
            numero_jugadores = %s,
            observaciones = %s,
            ruta_excel = %s
        WHERE id = %s
    """, (torneo_id, nombre, categoria_id, genero_id, tamano, numero_jugadores, observaciones, ruta_excel, cuadro_id))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(url="/admin/cuadros", status_code=303)
