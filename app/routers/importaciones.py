from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse
from openpyxl import load_workbook
from rapidfuzz import fuzz, process
import re
import unicodedata

from app.core import comprobar_admin, templates
from app.database import get_connection

router = APIRouter()

def normalizar_club_para_comparar(texto: str) -> str:
    if not texto:
        return ""

    texto = texto.upper().strip()

    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")

    texto = texto.replace("REAL CLUB DE TENIS", "")
    texto = texto.replace("REAL CLUB DE TENNIS", "")
    texto = texto.replace("CLUB DE TENNIS", "")
    texto = texto.replace("CLUB DE TENIS", "")
    texto = texto.replace("CLUB TENNIS", "")
    texto = texto.replace("CLUB TENIS", "")
    texto = texto.replace("CLUB ESPORTIU", "")
    texto = texto.replace("CLUB DEPORTIVO", "")
    texto = texto.replace("CLUB NATACIO", "")
    texto = texto.replace("CLUB NATACION", "")
    texto = texto.replace("CLUB", "")
    texto = texto.replace("TENNIS", "")
    texto = texto.replace("TENIS", "")

    texto = texto.replace(",", " ")
    texto = texto.replace("-", " ")

    texto = re.sub(r"[^A-Z0-9 ]", " ", texto)

    texto = re.sub(r"\bRCT\b", "", texto)
    texto = re.sub(r"\bCT\b", "", texto)
    texto = re.sub(r"\bTC\b", "", texto)
    texto = re.sub(r"\bRC\b", "", texto)
    texto = re.sub(r"\bCE\b", "", texto)
    texto = re.sub(r"\bC T\b", "", texto)

    texto = re.sub(r"\s+", " ", texto).strip()

    return texto



def separar_nombre_federacion(texto):
    if not texto or "," not in str(texto):
        return "", "", ""

    apellidos, nombre = str(texto).split(",", 1)

    nombre = nombre.strip().title()
    partes = apellidos.strip().title().split()

    apellido1 = partes[0] if len(partes) >= 1 else ""
    apellido2 = " ".join(partes[1:]) if len(partes) >= 2 else ""

    return nombre, apellido1, apellido2


def obtener_ano_nacimiento(fecha):
    if not fecha:
        return None

    if hasattr(fecha, "year"):
        return fecha.year

    texto = str(fecha).strip()

    if "/" in texto:
        return int(texto.split("/")[-1])

    return None


@router.post("/admin/importar-jugadores/excel-federacion")
async def importar_excel_federacion(
    archivo_excel: UploadFile = File(...),
    genero_id: int = Form(...),
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    wb = load_workbook(archivo_excel.file, data_only=True)
    ws = wb.active

    for fila in ws.iter_rows(min_row=2):
        licencia = fila[1].value
        nombre_completo = fila[2].value
        club = fila[9].value if len(fila) > 9 else ""
        fecha_nacimiento = fila[14].value if len(fila) > 14 else None

        if not licencia or not nombre_completo:
            continue

        nombre, apellido1, apellido2 = separar_nombre_federacion(nombre_completo)
        ano_nacimiento = obtener_ano_nacimiento(fecha_nacimiento)

        cur.execute("""
            INSERT INTO jugadores_importados
            (
                nombre,
                apellido1,
                apellido2,
                club,
                ano_nacimiento,
                numero_licencia,
                genero_id
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            nombre,
            apellido1,
            apellido2,
            club,
            ano_nacimiento,
            str(licencia).strip(),
            genero_id
        ))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(url="/admin/importar-jugadores", status_code=303)



@router.get("/admin/importar-jugadores")
def ver_importados(
    request: Request,
    buscar: str = "",
    ordenar: str = "apellido",
    genero_id: int = 0,
    admin: str = Depends(comprobar_admin)
):
    if ordenar == "club":
        order_by = "club, apellido1, apellido2, nombre"
    elif ordenar == "licencia":
        order_by = "NULLIF(numero_licencia, '') NULLS LAST, apellido1, apellido2, nombre"
    elif ordenar == "genero":
        order_by = "genero_id, apellido1, apellido2, nombre"
    else:
        order_by = "apellido1, apellido2, nombre"

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, nombre
        FROM generos
        ORDER BY id
    """)
    generos = cur.fetchall()

    condiciones = []
    parametros = []

    if buscar:
        texto_busqueda = f"%{buscar.strip()}%"
        condiciones.append("""
            (
                nombre ILIKE %s
                OR apellido1 ILIKE %s
                OR COALESCE(apellido2, '') ILIKE %s
                OR club ILIKE %s
                OR COALESCE(numero_licencia, '') ILIKE %s
            )
        """)
        parametros.extend([
            texto_busqueda,
            texto_busqueda,
            texto_busqueda,
            texto_busqueda,
            texto_busqueda
        ])

    if genero_id != 0:
        condiciones.append("genero_id = %s")
        parametros.append(genero_id)

    where_sql = ""

    if condiciones:
        where_sql = "WHERE " + " AND ".join(condiciones)

    cur.execute(f"""
        SELECT
            id,
            nombre,
            apellido1,
            apellido2,
            club,
            ano_nacimiento,
            numero_licencia,
            genero_id
        FROM jugadores_importados
        {where_sql}
        ORDER BY {order_by}
    """, parametros)

    jugadores = cur.fetchall()

    cur.close()
    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="importar_jugadores.html",
        context={
            "request": request,
            "jugadores": jugadores,
            "buscar": buscar,
            "ordenar": ordenar,
            "genero_id": genero_id,
            "generos": generos
        }
    )

def corregir_club_alias_manual(club: str) -> str:
    if not club:
        return club

    clave = normalizar_club_para_comparar(club)

    alias = {
        "ANDRES GIMENO": "Andrés Gimeno, CT",
        "BARA": "Barà, CT",
        "BLANES": "Blanes, CT",
        "CARDEDEU": "Cardedeu, CT",
        "DARO": "D'Aro, CT",
        "EL ROMANI": "El Romaní, CT",
        "LLAFRANC": "Llafranc, CT",
        "MANLLEU": "Manlleu, CT",
        "MATARO": "Mataró, CT",
        "MONTBLANC": "Montblanc, CT",
        "OLESA": "Olesa, CT",
        "PINEDA GAVA": "Pineda Gavà, CT",
        "RIPOLLET": "Ripollet, CT",
        "ROSES": "Roses, CT",
        "SABADELL": "Sabadell, CT",
        "SANT CELONI": "Sant Celoni, CT",
        "SERRAMAR": "Serramar, CT",
        "TARREGA": "Tàrrega, CT",
        "CERDANYOLA": "Cerdanyola, CT",
        "GIRONA": "Girona, CT",
        "BISBAL EMPORDA": "La Bisbal d'Empordà, CT",
        "PALLEJA": "Pallejà, CT",
        "SANT BOI": "Sant Boi, CT",
        "URGELL": "Urgell, CT",
        "COSTA BRAVA": "Costa Brava, CT",
        "ELS GORCHS": "Els Gorchs, CT",
        "POBLA CLARAMUNT": "La Pobla de Claramunt, CT",
        "VIC": "Vic, CT",
        "VILANOVA": "Vilanova, CT",
        "BARCINO": "Barcino, CT",
        "SALUT 1902": "La Salut 1902, CT",
        "FIGUERES": "Figueres, CT",
        "METLLA MAR": "L'Ametlla de Mar, CT",
        "LLEIDA": "Lleida, CT",
        "NATACIO SANT CUGAT": "Natació Sant Cugat, CT",
        "PINEDA": "Pineda, CT",
        "PORQUERES": "Porqueres, CT",
        "REUS MONTEROLS": "Reus Monterols, CT",
        "SITGES": "Sitges, CT",
        "TARRAGONA": "Tarragona, CT",
        "TORELLO": "Torelló, CT",
        "TORREDEMBARRA": "Torredembarra, CT",
        "TORTOSA": "Tortosa, CT",
        "BERGA": "Berga, TC",
        "BADALONA": "Badalona, TC",
        "BARCELONA 1899": "Barcelona-1899, RCT",
        "POLO": "Real Club de Polo",
        "TOPTEN": "Topten Tennis Club",
        "VALL HEBRON": "Vall d'Hebron",
        "EGARA": "Egara, Club",
        "FARNERS": "Farners Tennis Club",
        "FEDERACIO CATALANA": "Federació Catalana de Tennis",
    }

    return alias.get(clave, club)


@router.post("/admin/importar-jugadores/corregir-clubs")
def corregir_clubs_importados(admin: str = Depends(comprobar_admin)):
    conn = get_connection()
    cur = conn.cursor()

    # Clubs distintos que vienen del OCR
    cur.execute("""
        SELECT DISTINCT club
        FROM jugadores_importados
        WHERE club IS NOT NULL
          AND TRIM(club) <> ''
    """)
    clubs_importados = [fila[0] for fila in cur.fetchall()]

    # Clubs buenos desde tabla clubs
    cur.execute("""
        SELECT nombre
        FROM clubs
        WHERE nombre IS NOT NULL
          AND TRIM(nombre) <> ''
    """)
    clubs_tabla = [fila[0] for fila in cur.fetchall()]

    # Clubs buenos desde jugadores ya aprobados
    cur.execute("""
        SELECT DISTINCT club
        FROM jugadores
        WHERE club IS NOT NULL
          AND TRIM(club) <> ''
    """)
    clubs_jugadores = [fila[0] for fila in cur.fetchall()]

    clubs_buenos = sorted(set(clubs_tabla + clubs_jugadores))

    if not clubs_importados or not clubs_buenos:
        cur.close()
        conn.close()
        return RedirectResponse(
            url="/admin/importar-jugadores?ordenar=club",
            status_code=303
        )

    clubs_buenos_normalizados = {
        normalizar_club_para_comparar(club): club
        for club in clubs_buenos
    }

    for club_importado in clubs_importados:

        club_alias = corregir_club_alias_manual(club_importado)

        if club_alias != club_importado:
            cur.execute("""
                UPDATE jugadores_importados
                SET club = %s
                WHERE club = %s
            """, (club_alias, club_importado))
            continue

        club_importado_norm = normalizar_club_para_comparar(club_importado)
    
    

        mejor = process.extractOne(
            club_importado_norm,
            list(clubs_buenos_normalizados.keys()),
            scorer=fuzz.token_set_ratio
        )


        if not mejor:
            continue

        club_bueno_norm, puntuacion, _ = mejor
        club_bueno = clubs_buenos_normalizados[club_bueno_norm]

        print("CLUB IMPORTADO:", club_importado, "=>", club_importado_norm)
        print("MEJOR:", club_bueno, "=>", club_bueno_norm, "PUNTOS:", puntuacion)


        # Umbral alto para evitar cambios peligrosos.
        # Si corrige poco, luego bajamos a 85.
        if puntuacion >= 75 and club_importado != club_bueno:
            cur.execute("""
                UPDATE jugadores_importados
                SET club = %s
                WHERE club = %s
            """, (club_bueno, club_importado))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse(
        url="/admin/importar-jugadores?ordenar=club",
        status_code=303
    )

    
@router.get("/admin/importar-jugadores/aprobar/{jugador_id}")
def aprobar_jugador_importado(
    jugador_id: int,
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT nombre, apellido1, apellido2, club, ano_nacimiento, numero_licencia, genero_id
        FROM jugadores_importados
        WHERE id = %s   
    """, (jugador_id,))

    jugador = cur.fetchone()

    if jugador:
        cur.execute("""
            INSERT INTO clubs (nombre)
            VALUES (%s)
            ON CONFLICT (nombre) DO NOTHING
        """, (jugador[3],))
        
        cur.execute("""
            INSERT INTO jugadores
            (nombre, apellido1, apellido2, club, ano_nacimiento, numero_licencia, genero_id)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (numero_licencia) DO UPDATE
            SET genero_id = COALESCE(EXCLUDED.genero_id, jugadores.genero_id)
        """, jugador)

        cur.execute("""
            DELETE FROM jugadores_importados
            WHERE id = %s
        """, (jugador_id,)) 

        conn.commit()

    cur.close()
    conn.close()

    return RedirectResponse(url="/admin/importar-jugadores", status_code=303)

@router.post("/admin/importar-jugadores/borrar/{jugador_id}")
def borrar_importado(
    jugador_id: int,
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        DELETE FROM jugadores_importados
        WHERE id = %s
    """, (jugador_id,))

    conn.commit()

    cur.close()
    conn.close()

    return RedirectResponse(url="/admin/importar-jugadores", status_code=303)

@router.get("/admin/importar-jugadores/editar/{jugador_id}")
def editar_importado(
    request: Request,
    jugador_id: int,
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    # jugador
    cur.execute("""
        SELECT id, nombre, apellido1, apellido2, club, ano_nacimiento, numero_licencia, genero_id
        FROM jugadores_importados
        WHERE id = %s
    """, (jugador_id,))
    jugador = cur.fetchone()

    if not jugador:
        cur.close()
        conn.close()
        return RedirectResponse(url="/admin/importar-jugadores", status_code=303)

    # lista de clubs existentes
    cur.execute("""
    SELECT nombre
    FROM clubs
    ORDER BY nombre ASC
""")
    
    clubs = cur.fetchall()

    cur.execute("""
        SELECT id, nombre
        FROM generos
        ORDER BY id
    """)
    generos = cur.fetchall()

    cur.close()
    conn.close()

    return templates.TemplateResponse(
        request=request,
        name="editar_jugador_importado.html",
        context={
            "request": request,
            "jugador": jugador,
            "clubs": clubs,
            "generos": generos
        }
    )
@router.post("/admin/importar-jugadores/editar/{jugador_id}")
def guardar_importado(
    jugador_id: int,
    nombre: str = Form(...),
    apellido1: str = Form(...),
    apellido2: str = Form(""),
    club_existente: str = Form(""),
    club_nuevo: str = Form(""),
    ano_nacimiento: int = Form(...),
    numero_licencia: str = Form(...),
    genero_id: int = Form(...),
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    club_final = club_nuevo.strip() if club_nuevo.strip() else club_existente
    cur.execute("""
        INSERT INTO clubs (nombre)
        VALUES (%s)
        ON CONFLICT (nombre) DO NOTHING
    """, (club_final,))


    cur.execute("""
        UPDATE jugadores_importados
        SET nombre=%s, apellido1=%s, apellido2=%s,
            club=%s, ano_nacimiento=%s, numero_licencia=%s,
            genero_id=%s
        WHERE id=%s
    """, (
        nombre, apellido1, apellido2,
        club_final, ano_nacimiento, numero_licencia,
        genero_id,
        jugador_id
    ))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse("/admin/importar-jugadores", status_code=303)

@router.post("/admin/importar-jugadores/aprobar-seleccionados")
def aprobar_seleccionados(
    jugadores_ids: list[int] = Form(...),
    admin: str = Depends(comprobar_admin)
):
    conn = get_connection()
    cur = conn.cursor()

    for jugador_id in jugadores_ids:

        cur.execute("""
            SELECT nombre, apellido1, apellido2, club, ano_nacimiento, numero_licencia, genero_id
            FROM jugadores_importados
            WHERE id = %s
        """, (jugador_id,))

        jugador = cur.fetchone()

        if jugador:
            # Insertar club
            cur.execute("""
                INSERT INTO clubs (nombre)
                VALUES (%s)
                ON CONFLICT (nombre) DO NOTHING
            """, (jugador[3],))

            # Insertar jugador
            cur.execute("""
                INSERT INTO jugadores
                (nombre, apellido1, apellido2, club, ano_nacimiento, numero_licencia, genero_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (numero_licencia) DO UPDATE
                SET genero_id = COALESCE(EXCLUDED.genero_id, jugadores.genero_id)
            """, jugador)

            # Borrar importado
            cur.execute("""
                DELETE FROM jugadores_importados
                WHERE id = %s
            """, (jugador_id,))

    conn.commit()
    cur.close()
    conn.close()

    return RedirectResponse("/admin/importar-jugadores", status_code=303)
