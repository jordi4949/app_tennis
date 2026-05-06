from openpyxl import load_workbook
from app.database import get_connection

ruta = r"C:\Users\jordi\Desktop\app_tennis\Excels\copa_catalunya_1\cuadro_1.xlsx"

wb = load_workbook(ruta)
ws = wb.active

conn = get_connection()
cur = conn.cursor()

print("---- COMPARACIÓN CON JUGADORES OFICIALES ----")

for fila in ws.iter_rows(min_row=2):
    licencia = fila[1].value  # columna B
    nombre_excel = fila[2].value  # columna C

    if not licencia:
        continue

    licencia = str(licencia).strip()

    cur.execute("""
        SELECT id, nombre, apellido1, apellido2, numero_licencia
        FROM jugadores
        WHERE numero_licencia = %s
    """, (licencia,))

    jugador = cur.fetchone()

    if jugador:
        jugador_id, nombre, apellido1, apellido2, numero_licencia = jugador
        nombre_oficial = f"{nombre} {apellido1} {apellido2 or ''}".strip()
        print(f"✅ {licencia} | Excel: {nombre_excel} | Oficial: {nombre_oficial} | ID: {jugador_id}")
    else:
        print(f"❌ {licencia} | Excel: {nombre_excel} | NO ENCONTRADO")

cur.close()
conn.close()